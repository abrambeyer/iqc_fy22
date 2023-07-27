# -*- coding: utf-8 -*-
"""
Created on Tue Aug 24 14:46:43 2021

@author: nm184423
"""

#FY22 update of helpers.py

# library for interacting with the web browser.  Requires downloading web drivers before use.
# Drivers can be found at:  https://www.seleniumhq.org/
# For this project, we are using Google Chrome because only IE and Chrome are Vizient-approved web browsers.  Do not
# use Firefox for this project because there may be unexpected complications downloading files from Vizient's website.
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import NoSuchElementException
import math
import pandas as pd #libraries for interacting with dataframes in python
import numpy as np

from openpyxl import load_workbook #library for manipulating Excel files.
import openpyxl
import pyodbc #used to interact with the database
import os #used to help navigate and work with the operating system
from itertools import islice
import glob
import time
import shutil
import re
import urllib
import pathlib
import maskpass   #for masking password input (in SPYDER only).
import getpass    #for masking password input (if run outside of SPYDER).

################################################################################################
## AUTHOR:  R. Abram Beyer
## DESCRIPTION:  Helpers file which contains bulk of functions used in Vizient Q&A web scraper.

## FY22 UPDATES:
## Refactoring helpers.py functions to remove excess, duplicated code, improving waiter functions
## to make sure we use as many implicit waits as possible.
## Breaking up the original helpers.py into task-specific files for easier navigation.
################################################################################################
## UPDATE LOG:


################################################################################################


################################################################################################

# List of measures to exclude for testing.  These are excluded because Vizient currently has
# the time period radio button disabled on these templates.  We need to wait for them to re-activate them.
#list_of_total_revisits_measures = ['Readmission - Cardiology','Readmission - CT Surgery','Readmission - CT Surgery','Readmission - Gastroenterology','Readmission - Medicine General','Readmission - Neurology','Readmission - Neurosurgery','Readmission - Oncology','Readmission - Ortho/Spine','Readmission - Pulmonary/Critical Care','Readmission - Solid Organ Transplant','Readmission - Surgery General','Readmission - Trauma','Readmission - Vascular Surgery','Excess Days  - Cardiology','Excess Days - CT','Excess Days - Gastroenterology','Excess Days - Medicine General','Excess Days - Neurology','Excess Days - Neurosurgery','Excess Days - Oncology','Excess Days - Ortho/Spine','Excess Days - Pulmonary/Critical Care','Excess Days - Solid Organ Transplant','Excess Days - Surgery General','Excess Days - Trauma','Excess Days - Vascular Surgery']

list_of_total_revisits_measures = ['Excess Days  - Cardiology', 'Excess Days - CT', 'Excess Days - Gastroenterology', 'Excess Days - Medicine General', 'Excess Days - Neurology', 'Excess Days - Neurosurgery', 'Excess Days - Oncology', 'Excess Days - Ortho/Spine', 'Excess Days - Pulmonary/Critical Care', 'Excess Days - Solid Organ Transplant', 'Excess Days - Surgery General', 'Excess Days - Trauma', 'Excess Days - Vascular Surgery', 'Readmission - CT Surgery', 'Readmission - Cardiology', 'Readmission - Gastroenterology', 'Readmission - Medicine General', 'Readmission - Neurology', 'Readmission - Neurosurgery', 'Readmission - Oncology', 'Readmission - Ortho/Spine', 'Readmission - Pulmonary/Critical Care', 'Readmission - Solid Organ Transplant', 'Readmission - Surgery General', 'Readmission - Trauma', 'Readmission - Vascular Surgery', 'Readmissions - Medical', 'Readmissions - Surgical']

list_of_measures_to_run_twice_num_denom = ['Adverse Drug Events Rate','% Early Transfers Out']


################################################################################################


def process_date_df_to_dict(date_df):
    
    '''
    Takes a NM_Performance date dataframe as input:
    1. converts date strings to Vizient date dropdown menu format
    2. ensures correct column datatypes
    3. converts dataframe into a dictionary for the web scraper program to use as a mapping device.
    
    '''
    
    # replace Clarity month format with Vizient month format
    date_df = date_df.replace('January', 'Jan')
    date_df = date_df.replace('February', 'Feb')
    date_df = date_df.replace('March', 'Mar')
    date_df = date_df.replace('April', 'Apr')
    date_df = date_df.replace('May', 'May')
    date_df = date_df.replace('June', 'Jun')
    date_df = date_df.replace('July', 'Jul')
    date_df = date_df.replace('August', 'Aug')
    date_df = date_df.replace('September', 'Sep')
    date_df = date_df.replace('October', 'Oct')
    date_df = date_df.replace('November', 'Nov')
    date_df = date_df.replace('December', 'Dec')

    # format the year variables and zip everything into a dictionary
    date_df['begin_year'] = date_df['begin_year'].astype(str)
    date_df['end_year'] = date_df['end_year'].astype(str)
    date_df['zipped_begin'] = list(zip(date_df.begin_month, date_df.begin_year))
    date_df['zipped_end'] = list(zip(date_df.end_month, date_df.end_year))
    date_df['zipped_dates'] = list(zip(date_df.zipped_begin, date_df.zipped_end))
    period_lookup_dict = pd.Series(date_df.zipped_dates.values,
                                   index=date_df.period_type.values).to_dict()
    
    return(period_lookup_dict)
    
    
################################################################################################   
    
def collect_core_scraper_inputs():
    
    '''
    Function opens core_scraper_function_input_instructions.xlsx to set 
    time period and filepath variables in the core_scraper_function()
    
    '''
    
    input_df = pd.DataFrame(pd.read_excel(r"P:\Datastore02\Analytics\230 Inpatient Quality Composite\data\core_scraper_input_instructions\core_scraper_function_input_instructions.xlsx",sheet_name="inputs",engine='openpyxl'),columns=['input_item_DO_NOT_CHANGE','instructions_notes_DO_NOT_CHANGE','value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH'])
    
    input_dict = input_df[['input_item_DO_NOT_CHANGE','value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH']].set_index('input_item_DO_NOT_CHANGE').to_dict()
    
    return(input_dict) 
    
    
#################################################################################################   


def select_period_lookup_sql_query(select_num = 2):
    
    while select_num not in [1,2]:
        
        select_num = int(input('Selected number is not 1 or 2.  Please enter 1 or 2.'))
    
    '''
    returns a sql query string variable based on a number selection input for the 
    build_period_lookup_dict() function.
    
    '''
    
    if select_num == 1:
        sql_var = """
            SELECT
            period_month_begin.period_type
            ,start_dd2.MONTH_NAME as [begin_month]
            ,start_dd2.YEAR as [begin_year]
            ,end_dd2.MONTH_NAME as [end_month]
            ,end_dd2.YEAR as [end_year]
            FROM
            (
            SELECT
            pl.period_type
            ,start_dd.MONTH_BEGIN_DT as start_month_begin
            ,end_dd.MONTH_BEGIN_DT as end_month_begin
            from
            NM_Performance.period.period_lookup as pl
            left join clarity.dbo.DATE_DIMENSION as start_dd
            on cast(start_dd.CALENDAR_DT as date) = cast(pl.start_dts as date)
            left join clarity.dbo.DATE_DIMENSION as end_dd
            on cast(end_dd.CALENDAR_DT as date) = cast(pl.end_dts as date)
            WHERE
            pl.period_type in ('MONTH','fscl_ytd','fscl_qtd')
            and pl.end_dts = ?
            ) period_month_begin
            left join clarity.dbo.DATE_DIMENSION as start_dd2
            on cast(start_dd2.CALENDAR_DT as date) = cast(dateadd(mm,-1,period_month_begin.start_month_begin) as date)
            left join clarity.dbo.DATE_DIMENSION as end_dd2
            on cast(end_dd2.CALENDAR_DT as date) = cast(dateadd(mm,-1,period_month_begin.end_month_begin) as date)

            """
        return(sql_var)
    
    elif select_num == 2:
        sql_var = """
            DECLARE @begin_dts as date; SET @begin_dts = ?;
            DECLARE @end_dts as date; SET @end_dts = ?;


            SELECT
            bd.period_type
            ,start_dd2.MONTH_NAME as [begin_month]
            ,start_dd2.YEAR as [begin_year]
            ,end_dd2.MONTH_NAME as [end_month]
            ,end_dd2.YEAR as [end_year]
            FROM
            (
            SELECT
            'CUSTOM' as period_type
            ,start_dd.MONTH_BEGIN_DT as start_month_begin
            FROM
            clarity.dbo.DATE_DIMENSION as start_dd
            WHERE
            cast(start_dd.CALENDAR_DT as date) = cast(@begin_dts as date)
            ) bd
            LEFT JOIN
            (
            SELECT
            'CUSTOM' as period_type
            ,start_dd.MONTH_BEGIN_DT as end_month_begin
            FROM
            clarity.dbo.DATE_DIMENSION as start_dd
            WHERE
            cast(start_dd.CALENDAR_DT as date) = cast(@end_dts as date)
            ) ed
            on bd.period_type = ed.period_type
            left join clarity.dbo.DATE_DIMENSION as start_dd2
            on cast(start_dd2.CALENDAR_DT as date) = cast(bd.start_month_begin as date)
            left join clarity.dbo.DATE_DIMENSION as end_dd2
            on cast(end_dd2.CALENDAR_DT as date) = cast(ed.end_month_begin as date)
            """
        return(sql_var)  


#################################################################################################
        
        
def create_edw_connection():
    
    '''
    function tries to create a connection to MSSQL ODBC Node A.
    returns connection variable.
    
    If fails, prints an error message.
    
    '''
    try:
        # Establish a connection with Node A
        conn = pyodbc.connect('Driver={SQL Server};'
                                  'Server=edw00pd05wva.corp.nm.org\EDWIDS1;'
                                  'Database=clarity;'
                                  'Trusted_Connection=yes;')
        return(conn)
    except:
        print('Error connecting to the database.')





################################################################################################
    
def build_period_lookup_dict(input_dict = {},input_dict_key_list=None,time_period_choice_num=None):
    
    '''
    Function used to create a mapping time period dictionary used by the webscraper to properly 
    full report time periods.
    
    Function an input dictionary as input.  If the dictionary length == 0, Asks the users to choose between 3 different time period options.  Option 1 and 2 query
    the NM_Performance.period.period_lookup table to accurately find beginning and end dates of different time periods.
    
    time period choice refers to the different ways various reports need to pull data.  Importantly, READM/EDAC reports
    can only use quarters as input so option 3 would be the correct choice for READM/EDAC metrics.
    
    Otherwise, will use the input dictionary to create the time periods.
    
    I call this function 3 times in the core_scraper_function in order to create separate time period options for various
    Q&A template report types.  In general, THK reports require a 2-month lag on Performance Close month.  READM/EDAC also
    require a 2-month lag on Performance Close month but can only take quarters as input.  All other reports
    can be pulled using a 1-month lag.  This is due to the nature of Vizient's CDB report options.
    
    '''
    
    if input_dict_key_list is None:
        
        #establish connection to Node A.
        conn = create_edw_connection()
        
        if conn is None:
            return()

        time_period_choice = input(
            "Do you want the standard NM_Performance (choose 1) time periods or a custom date range (choose 2) or Vizient Q&A Quarter (choose 3)?")

        if time_period_choice == '1':

            # input the month end date in a format the NM_Performance.period.period_lookup table will accept
            end_dts = input("Enter the Performance Close Month end datetime (format:  'xx-xx-xxxx 23:59:59'')")

            # build the query
            sql_option1 = select_period_lookup_sql_query(select_num = 1)
            
            # Query the database and store the results of the query to a pandas dataframe
            try:
                test_period_data = pd.DataFrame(pd.read_sql(sql = sql_option1, con = conn, params = [end_dts]))
            except:
                print('Query error.')

            #clean up and convert df to dict    
            period_lookup_dict =  process_date_df_to_dict(test_period_data)

            return (period_lookup_dict)
        
        if time_period_choice == '2':
            
            #assign begin/end dates
            begin_dts = input("Enter the Custom beginning datetime (format:  'xx-01-xxxx')")
            end_dts = input("Enter the Custom Month end datetime (format:  'xx-xx-xxxx 23:59:59')")
            
            # build the query
            sql_option2 = select_period_lookup_sql_query(select_num = 2)
            
            # Query the database and store the results of the query to a pandas dataframe
            try:
                test_period_data = pd.DataFrame(pd.read_sql(sql = sql_option2, con = conn, params = [begin_dts, end_dts]))
            except:
                print('Query error.')


            period_lookup_dict =  process_date_df_to_dict(test_period_data)

            return (period_lookup_dict)
        if time_period_choice == '3':
            test_period_data2 = pd.DataFrame(columns=['period_type', 'qtrs'])

            answer = 'Y'
            qtr_data = ['QUARTER']
            while answer == 'Y':
                answer = input("Do you want to add another Fiscal Year Quarter?  (Y/N)").upper()
                while answer not in ['Y','N']:
                    print('That is not "Y" or "N".  Please fix your answer.')
                    answer = input("Do you want to add another Fiscal Year Quarter?  (Y/N)").upper()
                if answer == 'Y':
                    qtr_dts = input("Enter the beginning Vizient Quarter (format:  '20XX Quarter #')")
                    qtr_data.append(qtr_dts)
                elif 'N':
                    pass

            test_period_data = pd.DataFrame(qtr_data)

            value_list = [test_period_data[0].loc[0],np.array(test_period_data[0].loc[1:])]
            test_period_data2.loc[0] = value_list

            period_lookup_dict = pd.Series(test_period_data2.qtrs.values,index=test_period_data2.period_type.values).to_dict()
            return (period_lookup_dict)  
        
        
    else:
        
        #establish connection to Node A.
        conn = create_edw_connection()
        
        if conn is None:
            return()

        time_period_choice = time_period_choice_num

        if time_period_choice_num == '1':
            
            

            # input the month end date in a format the NM_Performance.period.period_lookup table will accept
            end_dts = input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH'][input_dict_key_list[len(input_dict_key_list) - 1]] 

            # build the query
            sql_option1 = select_period_lookup_sql_query(select_num = 1)
            
            # Query the database and store the results of the query to a pandas dataframe
            try:
                test_period_data = pd.DataFrame(pd.read_sql(sql = sql_option1, con = conn, params = [end_dts]))
            except:
                print('Query error.')

            #clean up and convert df to dict    
            period_lookup_dict =  process_date_df_to_dict(test_period_data)

            return (period_lookup_dict)
        if time_period_choice_num == '2':

            #assign begin/end dates
            begin_dts = input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH'][input_dict_key_list[0]] 
            end_dts = input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH'][input_dict_key_list[len(input_dict_key_list) - 1]] 
            
            # build the query
            sql_option2 = select_period_lookup_sql_query(select_num = 2)
            
            # Query the database and store the results of the query to a pandas dataframe
            try:
                test_period_data = pd.DataFrame(pd.read_sql(sql = sql_option2, con = conn, params = [begin_dts, end_dts]))
            except:
                print('Query error.')


            period_lookup_dict =  process_date_df_to_dict(test_period_data)

            return (period_lookup_dict)
        if time_period_choice_num == '3':
            
            test_period_data2 = pd.DataFrame(columns=['period_type', 'qtrs'])

            qtr_data = ['QUARTER'] + [item.strip() for item in input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH'][input_dict_key_list[0]].split(',')]
            
            test_period_data = pd.DataFrame(qtr_data)

            value_list = [test_period_data[0].loc[0],np.array(test_period_data[0].loc[1:])]
            test_period_data2.loc[0] = value_list
            period_lookup_dict = pd.Series(test_period_data2.qtrs.values,index=test_period_data2.period_type.values).to_dict()
            return (period_lookup_dict)  




################################################################################################
        
def clean_cohort_worksheet_obj(worksheet_obj, scenario_param = 1):
    
    '''
    Function used within gather_cohort_data()
    
    Function takes an openpyxl worksheet object as input.  
    Cleans up the values and puts the value matrix into a pandas dataframe.
    isolates the Hospital and Cohort columns 
    The cleans up the cohort name column
    
    scenario_param controls which type of cleaning the function should do 
    based off the Vizient cohort file year and type.
    
    '''
    
    if scenario_param == 1:
        
        data = worksheet_obj.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)
        df['Hospital'] = df.index
        df = pd.DataFrame(df, columns=['Hospital', 'Cohort'])
        #They added a comma in LSCCMC cohort name in the 2020 calculators....
        df['Cohort'] = df['Cohort'].str.replace(',','').str.replace('.','')
        return(df)
    
    elif scenario_param == 2:
        
        # open and process cohort worksheet
        # Starting with FY2020, there is a Medicare ID column and Short name column instead of
        # concatenated 'Hospital' column
        
        data = worksheet_obj.values
        cols = next(data)[0:]
        data = list(data)
        # recreate the Hospital column from FY19.  Just concatenate the medicare ID to the Short name
        data[:] = [list(i) for i in data]
        [i.insert(0, str(i[0]) + ' ' + str(i[1])) for i in data]
        data[:] = [tuple(i) for i in data]
        idx = [r[0] for r in data]
        data = (islice(r, 0, None) for r in data)
        cols = list(cols)
        cols.insert(0, 'Hospital')
        cols = tuple(cols)
        df = pd.DataFrame(data, index=idx, columns=cols)
        df['Hospital'] = df.index
        df = pd.DataFrame(df, columns=['Hospital', 'Cohort'])
        
        return(df)
    
    elif scenario_param == 3:
        
        data_crit = worksheet_obj.values
        # skip the header row
        cols_crit = next(data_crit)
        cols_crit = next(data_crit)
        data_crit = list(data_crit)
        # recreate the Hospital column from FY19.  Just concatenate the medicare ID to the Short name
        data_crit[:] = [list(i) for i in data_crit]
        [i.insert(0, str(i[0]) + ' ' + str(i[1])) for i in data_crit]
        data_crit[:] = [tuple(i) for i in data_crit]
        idx_crit = [r[0] for r in data_crit]
        data_crit = (islice(r, 0, None) for r in data_crit)
        cols_crit = list(cols_crit)
        cols_crit.insert(0, 'Hospital')
        cols_crit = tuple(cols_crit)
        df_crit = pd.DataFrame(data_crit, index=idx_crit, columns=cols_crit)
        df_crit['Hospital'] = df_crit.index
        # no cohort column so we need to fill this in.
        df_crit['Cohort'] = 'Critical Access & Small Community'
        df_crit = pd.DataFrame(df_crit, columns=['Hospital', 'Cohort'])
        
        return(df_crit)

################################################################################################
        
#function to create a pandas dataframe from provide cohort excel files.
def gather_cohort_data(input_dict_string = 0):
    
    '''
    Function takes no input
    
    Iterates over a folder with an excel file like QA_Calculator_..._Cohorts.xlsx
    
    The function is designed to handle each of Vizient's release period cohort excel files and
    
    extract the hospital name and cohort assignment for that period. 
    
    function returns a pandas dataframe consisting of hospital names and cohort names.
    
    Vizient cohort file can be found under Report Express on the Vizient website.
    
    Each Q&A period cohort file is stored within the 230 Inpatient Quality Composite folder.  
    
    When this function asks for the working directory, the input should be the folder which contains
    
    the current baseline period Q&A calculator cohort.xlsx file (stored within the 230 Inpatient Quality Composite workd folder)
    
    '''
    
    
    try:
        
        if input_dict_string == 0:
            
            wd = input('Enter file path for Vizient cohort file (File Should be like this: QA_Calculator_..._Cohorts.xlsx).')
        
        else:
            
            wd = input_dict_string
            
        # change directory to directory with file.  abspath function normalizes the directory path.
        os.chdir(os.path.abspath(wd))
        # print(os.listdir())
    except:
        print('Something is wrong with cohort file path.')
        return
    try:
        # find the cohort excel file.  The file name should have QA_ and _Cohorts.xlsx in the name string.
        cohort_file_list = [file for file in os.listdir(wd) if
                            ('QA_' in file and '_Cohorts.xlsx' in file) and not (file.startswith('~'))]
        
        try:
            # There should only be one file in the cohort directory
            if len(cohort_file_list) > 1:
                print('More than one cohort list file.  There should only be one.')
                return
            # There should be at least one cohort file in the directory
            elif len(cohort_file_list) == 0:
                print('No cohort list files in this directory.  Check for empty directory or check file name.')
                return
            else:
                # Once you get one cohort excel file found, load it.
                try:
                    wb = openpyxl.load_workbook(filename=cohort_file_list[0], data_only=True, read_only=True)
                    
                except:
                    print('Could not load workbook')
                    return
                

                # once you have the workbook loaded, find all the period cohort worksheets.
                # Unfortunately, the sheet names change each period and the critical access cohort is on a separate
                # sheet starting in FY2020 Period 1 so we don't have one constant worksheet name to call so we must
                # find them.
                # find the current period cohort worksheet name
                cohort_sheet_list = [sheet for sheet in wb.sheetnames if
                                     ('Q&A' in sheet and 'Cohorts' in sheet) or (sheet == 'Crit Acc. & Small Comm')]
                
                # if there is only one sheet, it likely lacks the critical access sheet.  Probably FY19.
                if len(cohort_sheet_list) == 1:
                    try:
                        ws = wb[cohort_sheet_list[0]]
                    except:
                        print('Something went wrong opening the worksheet.')
                        return
                    
                    #clean up the cohort worksheet
                    df = clean_cohort_worksheet_obj(ws, scenario_param=1)
                    return (df)
                elif len(cohort_sheet_list) == 0:
                    # empty list.  failed to find any worksheets.  return and figure out the problem.
                    # if there are
                    print('Did not find any worksheets.')
                    return
                elif len(cohort_sheet_list) == 2:
                    cohort_sheet_list_index = 0
                    crit_access_list_index = 0
                    for i, item in enumerate(cohort_sheet_list):
                        if ('Q&A' in item and 'Cohorts' in item):
                            cohort_sheet_list_index = i
                        if item == 'Crit Acc. & Small Comm':
                            crit_access_list_index = i

                    # open and process cohort worksheet
                    try:
                        ws = wb[cohort_sheet_list[cohort_sheet_list_index]]
                    except:
                        print('Something went wrong opening the worksheet.')
                        return

                    data = ws.values
                    columns = list(next(data)[0:])

                    # in order to handle the 2019 Period 4 calculator, look for 'Hospital' column name as indicator
                    if 'Hospital' in columns:
                        
                        #clean up the cohort worksheet
                        df = clean_cohort_worksheet_obj(ws, scenario_param=1)
                        
                        return (df)
                    else:
                        
                        
                        df = clean_cohort_worksheet_obj(ws, scenario_param=2)

                        # now do almost the same for the critical access worksheet

                        try:
                            ws_crit = wb[cohort_sheet_list[crit_access_list_index]]
                        except:
                            print('Something went wrong opening the worksheet.')
                            return
                        
                        
                        #slightly different cleaning procedure for critical access FY2020
                        df_crit = clean_cohort_worksheet_obj(ws_crit, scenario_param=3)
                        
                        
                        df_list = [df, df_crit]
                        df_final = pd.concat(df_list)
                        # They added a comma in LSCCMC cohort name in the 2020 calculators....
                        df_final['Cohort'] = df_final['Cohort'].str.replace(',','').str.replace('.','')
                        return (df_final)
                        
        

        except:
            print('Something went wrong finding files ending in _Cohorts.xlsx.')
            return
    except:
        print('No cohort file found.')
        return
  
        


################################################################################################
        
    
def gather_hyperlink_files(input_dict_string = 0):
    
    '''
    Function takes no input.
    
    Function asks for a filepath as input.  Filepath should be the folder location of the baseline period template_hyperlink
    excel files.  Template hyperlink files are stored in 230 Inpatient Quality Composite folder/data/hyperlink data folder.
    
    Function gathers all files ending in _links.xlsx and returns a list of file names along with the filepath to their
    location.
    
    IMPORTANT NOTE:
    
    For example, in FY21, we used the 2020 Period 2 calculators as baseline.  Therefore, the baseline template_hyperlinks
    from 230 Inpatient Quality Composite/data/hyperlink data/2020/period2_template_hyperlinks should be used.
    
    The core_scraper_function function will modify these files as it iterates.  This function is designed to 
    loop over each report template hyperlink, open the hyperlink in the web browser and download the report.  The LAST STEP
    WILL DELETE THE ROW FROM THE EXCEL FILE.  I designed the function this way so if the web scraper stops or errors out
    midway, it can pick up where it left off.
    
    THEREFORE, YOU SHOULD ALWAYS COPY THE BASELINE report template files before running the core_scraper_function() function.
    I always make a copy of the baseline hyperlink files into a new folder called 001_current hyperlink files for scraping.
    
    This way the copies will be modified and not the originals.
    
    '''
    
    try:
        if input_dict_string == 0:
            
            wd = input('Enter file path for hospital hyperlink folder (template_hyperlinks).')
            
        else:
            wd = input_dict_string
        # change directory to directory with file.  abspath function normalizes the directory path.
        os.chdir(os.path.abspath(wd))
    except:
        print('Something is wrong with Vizient hyperlink excel file path.')
        return
    files = os.listdir(os.curdir)
    # Filter folder files to only include '_links.xlsx' excel files.
    files = [ii for ii in files if '_links.xlsx' in ii and not ii.startswith('~')]
    return(files,wd)
    
    
################################################################################################



def get_report_template_links_orig(input_dict_string = 0):
    
    """
    This function first runs the gather_hyperlink_files() function to get a list of all _links.xlsx files along with their
    location filepath so they can be opened.  Appends each file together and removes nulls and zero values.  Basic dataframe
    cleaning.
    
    The function returns a cleaned dataframe of cdb report template hyperlinks and hospital names 
    for each NM hospital in the Vizient Q&A.  
    
    The purpose of this function is to create a dataframe of NM hospitals and CDB Vizient Q&A report templates
    to loop over so the web scraper can open each report and also understand which hospital & cohort the report
    belongs to.
    
    Function also returns filepath to the hyperlink files so they can be opened again later.
    
    IMPORTANT NOTE:
    
    For example, in FY21, we used the 2020 Period 2 calculators as baseline.  Therefore, the baseline template_hyperlinks
    from 230 Inpatient Quality Composite/data/hyperlink data/2020/period2_template_hyperlinks should be used.
    
    The core_scraper_function function will modify these files as it iterates.  This function is designed to 
    loop over each report template hyperlink, open the hyperlink in the web browser and download the report.  The LAST STEP
    WILL DELETE THE ROW FROM THE EXCEL FILE.  I designed the function this way so if the web scraper stops or errors out
    midway, it can pick up where it left off.
    
    THEREFORE, YOU SHOULD ALWAYS COPY THE BASELINE report template files before running the core_scraper_function() function.
    I always make a copy of the baseline hyperlink files into a new folder called 001_current hyperlink files for scraping.
    
    This way the copies will be modified and not the originals.
    
    """
    
    #find the folder with Vizient calculator template hyperlinks and put file names in a list
    try:
        file_names = gather_hyperlink_files(input_dict_string)
    except:
        print('Problem gathering hyperlink files.')
        return
    #Get all unique report templates for all hospitals
    #initialize empty dataframe to store hyperlinks
    hyperlinks = pd.DataFrame()
    #iterate through list of hyper link files obtain from Vizient calculators and store measure name & hyperlink.
    for ii, item in enumerate(file_names[0]):
        ##UL007
        dataframe_ob = pd.DataFrame(pd.read_excel(item,sheet_name="Sheet1",engine='openpyxl'),columns=['Hospital','Formal Name','Hyperlink','JobStoreID','ReportID','AdjustmentModel','AHRQ Version','Keyword/Metric','Domain'])
        hyperlinks = hyperlinks.append(dataframe_ob)
        
    #clean up the resulting dataframe
         
    #Remove any zero rows or null rows.  These are measures without a report template link.
    hyperlinks = hyperlinks[(hyperlinks['Hyperlink'].notnull()) & (hyperlinks['Hyperlink'] != 0)]
    
    #Convert JobStoreID and ReportID back to integer to remove decimal point.
    hyperlinks[["JobStoreID", "ReportID"]] = hyperlinks[["JobStoreID", "ReportID"]].astype(int)
    #Remove duplicates.  Only require a unique list of report templates for all hospitals
    hyperlinks = hyperlinks.drop_duplicates()
    
    #Keyword/Metric for VWH only is not all caps...for some reason so I now have to convert this column to all caps.
    hyperlinks['Keyword/Metric'] = hyperlinks['Keyword/Metric'].str.upper()
    return(hyperlinks,file_names[1])
    
    
    
################################################################################################
    
    
    
def create_hyperlink_dict(merged_df):
    
    '''
    Functions takes a dataframe as input containing cohort name, 
    hospital name and Q&A cdb report template hyperlinks
    and returns a dictionary.
    '''
    
    
    merged_df = pd.DataFrame(merged_df,columns=['Cohort','Formal Name','Hyperlink','JobStoreID','ReportID','AdjustmentModel','AHRQ Version','Keyword/Metric','Domain'])
    merged_df = merged_df.drop_duplicates()
    merged_df['zipped_data'] = list(zip(merged_df.Hyperlink,merged_df.JobStoreID,merged_df.ReportID,merged_df.Cohort,merged_df['Formal Name'],merged_df['AdjustmentModel'],merged_df['AHRQ Version'],merged_df['Keyword/Metric'],merged_df['Domain']))
    merged_df['zipped_keys'] = list(zip(merged_df.Cohort,merged_df['Formal Name']))
    lookup_data_container = pd.Series(merged_df.zipped_data.values,index=merged_df.zipped_keys.values).to_dict()
    return(lookup_data_container)
    
    
################################################################################################


def check_and_make_dir(dest_file_path):
    
    '''
    Takes a filepath as input.  Checks if the filepath already exists.  
    If it does not exist, create it.  If it does, pass.
    
    '''
    
    try:
        # check if folder already exists.  If it does not exist, create it.
        if os.path.isfile(dest_file_path) == False:
            os.mkdir(dest_file_path)
    except:
        pass
    
    
    
 ################################################################################################


# function takes the hyperlink dictionary created above and generates a folder structure to store the files.
def create_folder_structure(links_dict, input_dict_string = 0):
    
    '''
    Takes the cohort, hospital, hyperlink lookup dictionary as input.
    Asks for a filepath location where you would like to create a new folder structure to hold
    all sorted, named output files from the Q&A CDB report templates.
    
    Names the main folder 'Vizient Q&A Files,'then creates sub folders named after each Vizient Q&A cohort for the 
    given Q&A year that is present in the cohort/hyperlink dictionary.  This is controlled by the template_hyperlinks files.
    If a NM hospital is present in one of those files, its cohort will be included in the resulting lookup dictionary.
    
    For example, NMH is the only AMC hospital.  If we include the nmh_links.xlsx file, NMH's hyperlinks will be included 
    as well as the AMC cohort.  However, since NMH is the only AMC hospital, if we exclude NMH, we will also exclude AMCs
    and not recreate an AMC subfolder in this folder structure.
    
    Returns the filepath destination of the new folder structure.
    
    '''
    
    # create main folder
    if input_dict_string == 0:
        new_dir_path = input('Enter path of location where you want to store the files.')
    else:
        print('getting folder loc from excel file...')
        new_dir_path = input_dict_string
    folder_name = 'Vizient Q&A Files'
    new_dir_path = os.path.join(os.path.abspath(new_dir_path), folder_name)
    
    check_and_make_dir(new_dir_path)
    
    # create sub folders per cohort
    # iterate over dictionary keys and create a distinct list of cohort names
    distinct_cohort_names_list = list(set([i[0] for i in links_dict.keys()]))
    
    for i, item in enumerate(distinct_cohort_names_list):
        
        check_and_make_dir(os.path.join(new_dir_path, item))
        
    return (new_dir_path)
    
    
################################################################################################    
    
def implicitly_wait_select_click_by_elem_type(driver, elem_obj, elem_type = By.XPATH):
    
    '''
    Function takes selenium driver variable and (xpath, id, class, etc.) as input.
    implicitly waits until the element is present.  Then tries to click the element.
    if this click fails, try to find the element again and try clicking again.
    
    Returns the driver variable and DOM element variable.
    
    ''' 
    #ignore some common warnings
    ignored_exceptions=(NoSuchElementException,StaleElementReferenceException,)
    #find the DOM element and implicitly wait until it is present on the DOM.
    #hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((elem_type, elem_obj)))
    
    for i in range(5):
        try:
            hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((elem_type, elem_obj)))
            hosp_element.click()
            return(driver,hosp_element)
        except:
            try:
                hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((elem_type, elem_obj)))
                hosp_element.click()
                return(driver,hosp_element)
            except:
                print('attempt:',i+1)
                time.sleep(3)
                pass
        #if the above fails, try one more time.
        #hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((elem_type, elem_obj)))
        #hosp_element.click()
    return(driver,'FAIL')
    

################################################################################################ 
    
def implicitly_wait_select_by_elem_type(driver, elem_obj, elem_type = By.XPATH):
    
    '''
    Function takes selenium driver variable and (xpath, id, class, etc.) as input.
    implicitly waits until the element is present.  Then tries to click the element.
    if this click fails, try to find the element again and try clicking again.
    
    Returns the driver variable and DOM element variable.
    
    ''' 
    #ignore some common warnings
    ignored_exceptions=(NoSuchElementException,StaleElementReferenceException,)
    #find the DOM element and implicitly wait until it is present on the DOM.
    
    for i in range(5):
        try:
            hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((elem_type, elem_obj)))
            return(driver,hosp_element)
        except:
            try:
                hosp_element = WebDriverWait(driver, 120,ignored_exceptions=ignored_exceptions).until(expected_conditions.presence_of_element_located((elem_type, elem_obj)))
                return(driver,hosp_element)
            except:
                print('attempt:',i+1)
                time.sleep(3)
                pass
        #if the above fails, try one more time.
    return(driver,'FAIL') 

    
################################################################################################   
    
def driver_setup():
    
    '''
    Function used to create the Selenium webdriver.
    Takes no input parameters and returns a customized selenium webdriver variable 
    
    '''
    #initiate Chrome options variable
    #options = webdriver.EdgeOptions()
    options = webdriver.ChromeOptions()
    #turn off "automation flag."  Vizient started blocking bots.  This option basically hides the 
    #fact that this is bot.
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    #Turn off "Save Password?" Chrome popup.  Chrome kept asking to save the password.  Stop it, Chrome.
    prefs = {"credentials_enable_service": False, "profile.password_manager_enabled" : False}
    options.add_experimental_option("prefs", prefs)
    #set the path of the Chrome driver
    current_path = r'P:\Datastore02\Analytics\230 Inpatient Quality Composite\python\fy22\inpatient_quality_composite_pipeline'
    #current_path = pathlib.Path(__file__).parent.absolute()
    #driver_path = r'drivers\msedgedriver.exe'
    driver_path = r'drivers\chromedriver.exe'
    #initiate the Chrome browser using the above options.
    browser = webdriver.Chrome(options = options,executable_path = os.path.abspath(os.path.join(current_path,driver_path)))
    
    return(browser)
    
    
################################################################################################   
    
    
# Function to login to Vizient and open up the browser.

def vizient_login(browser_obj):
    
    '''
    Function takes a selenium driver object as input.
    Opens www.vizientinc.com
    clicks the home page login button.  clicks the vizient member login button.
    enters the user email and password and clicks submit.
    returns a selenium driver object.
    
    '''
    
    # initialize Chrome
    
    
    #browser = webdriver.Edge(executable_path = os.path.abspath(os.path.join(current_path,driver_path)))
    browser_obj.get('https://www.vizientinc.com/')
    #browser.implicitly_wait(30)
    try:
        browser_obj.maximize_window()
    except:
        pass
    # Click the login button
    # If the screen is not maximized, it will provie the mobile buttons
    try:
        browser, login_btn = implicitly_wait_select_click_by_elem_type(browser_obj, elem_obj = '//*[@id="header-bar"]/div[2]/nav[3]/ul/li[2]/a/span', elem_type = By.XPATH)
        
    except:
        try:
            
            browser, mobile_login = implicitly_wait_select_click_by_elem_type(browser, '//*[@id="mobile-header"]/div/div/a[1]', elem_type = By.XPATH)
            
        except:
            print('LOGIN BUTTON: Was not able to find an element with that name.')

    # Click the Vizient Member Dashboard login
    try:
        
        browser, vizient_member_login_btn = implicitly_wait_select_click_by_elem_type(browser, '//*[@id="__next"]/div/main/section[2]/div/div/div[1]/div/article/div/a[1]', elem_type = By.XPATH)
        
    except:
        print('VIZENT MEMBER LOGIN BUTTON: Was not able to find an element with that name.')
    # time.sleep(0.4)

    # Enter login email
    try:
        
        
        browser, vizient_email_input = implicitly_wait_select_click_by_elem_type(browser, "username", elem_type = By.NAME)
        
        login_email = input('Enter Vizient Login email.')
        vizient_email_input.send_keys(login_email)
    except:
        print('EMAIL FORM: Was not able to find an element with that name.')

    # click submit
    
    browser, vizient_email_next_btn = implicitly_wait_select_click_by_elem_type(browser, "idp-discovery-submit", elem_type = By.ID)

    try:
        #enter password into password form
        browser, password_form = implicitly_wait_select_click_by_elem_type(browser, "okta-signin-password", elem_type = By.ID)
    
        
        if any(['SPYDER' in name.upper() for name in os.environ]):
            #getpass does not work in the Spyder IDE.  So, check for SPYDER.  If running in Spyder, use maskpass instead.
            login_password = maskpass.advpass(prompt = 'Enter Vizient login password.')
            
        else:
            login_password =  getpass.getpass(prompt='Enter Vizient login password.')
            
        #login_password = input('Enter Vizient login password.')
        password_form.send_keys(login_password)
    except:
        print('PASSWORD FORM: Was not able to find an element with that name.')
    
    
    #click submit
    try:
        browser, vizient_submit_password = implicitly_wait_select_click_by_elem_type(browser, "okta-signin-submit", elem_type = By.ID)
    except:
        print('CLICK SUBMIT: Was not able to find an element with that name.')
    browser_page_source = browser.page_source
    return (browser,browser_page_source)  
    
    
################################################################################################
    
def setup_webdriver_and_vizient_login():
    
    '''
    Function takes no input parameters.
    
    Runs two main functions:  driver_setup() and vizient_login()
    
    Then checks the webpage source for a change.  Then checks whether or not login was successful.
    
    If not successful, runs vizient_login() function once more.  If successful, moves on and returns 
    the selenium driver variable.
    
    '''
    
    #setup and initiate selenium webdriver variable
    webdriver_var = driver_setup()
    
    #start up Chrome, click login buttons, enter email and password.
    webdriver_output, source = vizient_login(webdriver_var)
    
    #Check if login failed for some reason.  If we had a type-o or accidently messed up during sign-in
    #we don't want to have to redo all the initial steps.  So, if we failed login, let me try again.
    #check whether or not browser page source has the phrase "Sign in failed!"  If so, something went wrong.  Try agin.
    
    #wait until page source changes to check whether sign-in failed or not.
    while source == webdriver_output.page_source:
        print('waiting for page source to change..')
        time.sleep(1)
    time.sleep(2)    
    #check if error message is present on the page source.    
    signin_fail_check = 'Sign in failed!' in webdriver_output.page_source
    

    if signin_fail_check == True:
        print("LOGIN FAILED.  TRY AGAIN.")
        #failed.  Run the vizient_login() function one more time.
        webdriver_var = vizient_login(webdriver_var)
        return(webdriver_var)
    #successful!  move on.
    else:
        print('LOGIN SUCCESSFUL.')
        return(webdriver_var)
    

    
################################################################################################ 

def grab_download_folder_dir():
    
    '''
    function tests 2 different filpaths to download folders.
    I created this function to test for the download folder location so 
    the script wouldn't fail.
    
    returns a filepath string variable.
    
    '''
    '''
    try:
        len(os.listdir(r'C:/Users/NM184797/Downloads'))
        download_folder_dir_var = 'C:/Users/NM184797/Downloads'
    except:
        len(os.listdir(r'H:\Downloads'))
        download_folder_dir_var = r'H:\Downloads'
    '''
    try:
        len(os.listdir(r'C:/Users/NM184797/Downloads'))
        download_folder_dir_var = 'C:/Users/NM184797/Downloads'
    except:
        len(os.listdir(r'H:\Downloads'))
        download_folder_dir_var = r'H:\Downloads'
    
    return(download_folder_dir_var)
        


################################################################################################  


def open_template_report(hyperlink,browser_var):
    '''
    takes a cdb report template hyperlink and selenium driver variable as input.
    opens the hyperlink in the browser and returns the driver variable.
    '''
    browser_var.get(hyperlink)
    return(browser_var)


################################################################################################  

def check_num_downloaded_files_open_template_max_screen_and_scroll(driver,download_folder_dir_var,hyperlink_var):
    
    '''
    takes a file path to the downloads folder, a selenium driver variable and cdb report template hyperlink string variable
    as input.
    Checks how many files are already downloaded in the download folder (used later to tell when a file has completed downloading)
    Opens the hyperlink in the browser
    Makes sure the screen is maximized (to avoid DOM element name changes due to screen size)
    scrolls to the Risk Adjustment div to ensure the top of the page is visible and clickable.
    
    '''
    
    #check how many files are already in the downloads folder.  This is used at the end of the loop
    #to know when a file has completed download and also to find the most recently downloaded file.
    num_already_downloaded_files = len(os.listdir(download_folder_dir_var))
    
    #open the cdb report template hyperlink
    driver = open_template_report(hyperlink_var, driver)
    
    #make sure the browser window is maximized
    try:
        
        driver.maximize_window()
    except:
        pass
    
    # scroll down
    driver, div_element1 = implicitly_wait_select_by_elem_type(driver, elem_obj = "//div[@id='divRiskAdjustment']", elem_type = By.XPATH)
    driver.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element1)
    
    return(driver,num_already_downloaded_files)

################################################################################################
    
#FY21 COVID FILTER FUNCTIONS

def find_edit_custom_list_link(driver):
    '''
    First check if "Edit Custom List" is present in the page source. If not, pass.  Otherwise, we need to remove
    the custom list.
    get a list of all instances of the text 'Edit Custom List.'  
    If that phrase is present in the HTML, then there has been a custom list created.  
    There could be several, but we care about the first one at the top of the page.
    '''
    time.sleep(0.5)
    if "Edit Custom List" in driver.page_source:
        
        
        
        edit_custom_list_link = driver.find_elements_by_xpath("//*[contains(text(),'" + 'Edit Custom List' + "')]")
        if edit_custom_list_link:
            return(edit_custom_list_link)
        else:
            return(edit_custom_list_link) 
    else:
        
        return([])
        

################################################################################################

def find_remove_custom_group_by_bttn(driver):
    
    '''
    After clicking the 'Edit Custom List' link, you need to click 'Remove Custom Group By'
    There are actually many of these buttons on the DOM, so we need to do a conditional click.
    If PSI measure, then click the second button id.  If not, then click the first.
    
    #Click the button with the value 'Remove Custom Group By' to remove the custom list.
    
    '''
    remove_custom_group_by_bttn = driver.find_elements_by_xpath("//*[@id = 'ctl00_ContentPlaceHolder1_ucctl11_Custom2_cmdRemoveCustomGroupby' or @id = 'ctl00_ContentPlaceHolder1_ucctl11_Custom_cmdRemoveCustomGroupby']")
    if remove_custom_group_by_bttn:
        return(remove_custom_group_by_bttn)
    else:
        return(False) 
        
        
################################################################################################
    
def find_alert_popup(driver):
    
    '''
    Takes a selenium driver variable as input.
    Switch browser command to an alert popup window.
    returns selenium driver variable.
    '''
    
    obj = driver.switch_to.alert
    if obj:
        return obj
    else:
        return False
    
################################################################################################


def remove_custom_covid_list_from_template(browser_var,measure_name_var):
    
    '''
    Takes a selenium driver variable, measure name variable as input (to check for PSI or not).
    Finds the "Edit Custom List" menu on the DOM.  If not present, then passes.
    If find this menu, then removes all custom lists.
    Returns the selenium driver variable.
    
    '''
    
    #give it a couple seconds to load
    time.sleep(1)
    #find the first appearance of text 'Edit Custom List.'  If it doesn't exist,
    #then there isn't this restriction.  READM, EDAC, THK don't have this restriction.
    edit_custom_list_list = find_edit_custom_list_link(browser_var)
    
    #some templates do not use a custom list.  Check to see if there is a custom list.  If not, pass.  If yes, then
    #click the link to open up the editor.
    if len(edit_custom_list_list) == 1:
       
        edit_custom_list_list[0].click()
        
    else:
        return(browser_var)
    
    #click Remove Custom Group By
    
    remove_custom_group_by_bttn = WebDriverWait(browser_var, 3).until(find_remove_custom_group_by_bttn)
    
    #if it's a PSI, then choose the second button id element.  Else, 
    if measure_name_var in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
        remove_custom_group_by_bttn[1].click()
        
    else:
        remove_custom_group_by_bttn[0].click()
        
        
    obj = WebDriverWait(browser_var, 3).until(find_alert_popup)

    obj.accept()
    
    return(browser_var) 
    

################################################################################################

def update_group_by_select_to_default(browser_var,measure_name_var):
    
    '''
    Finds multiple groupby dropdown menu element.
    If measure is a PSI, change multiple group by to AHRQ Safety.  Else change to or make sure Hospital / Hospital System is selected.
    '''
    
    #find and click multiple groupby div (dropdown menu)
    browser_var, multgroupbydiv = implicitly_wait_select_click_by_elem_type(browser_var, elem_obj = 'ctl00_ContentPlaceHolder1_GroupByOutcomesWithMult', elem_type = By.ID)
    
    #if psi, then select AHRQ, else Hospital/Hospital System
    if measure_name_var in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
        Select(browser_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdMeasuresByWithMult']")).select_by_visible_text('AHRQ Safety')
        time.sleep(0.2)
        multgroupbydiv.click()
        time.sleep(0.2)
    else:
        Select(browser_var.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdMeasuresByWithMult']")).select_by_visible_text('Hospital / Hospital System')
        time.sleep(0.2)
        multgroupbydiv.click()
        time.sleep(0.2)
    return(browser_var)


################################################################################################
    
def while_loop_handler_function(selenium_function):
    
    '''
    This function takes another selenium function as input.
    Will try the function 5 times and handle for stale elements on the dom.
    returns the selenium dom element variable.
    Using this function to make sure we try to find elements mutliple times before giving up.
    '''
    
    attempt_num = 1
    while (attempt_num != 0):
        try:
            dom_element = selenium_function
            attempt_num = 0
        except StaleElementReferenceException:
            if attempt_num == 4:
                raise
            attempt_num += 1
            time.sleep(0.5)
    return(dom_element)



################################################################################################

def remove_covid_19_gen_med_subservice_rest(browser_var,measure_name_var):
    
    '''
    this function is used to remove the covid 19 subservice line restriction on
    only the general medicine templates (but not readmission or edac for some reason...Just mort,los,dcost).
    It uses the 3 functions defined above.

    Critical access uses a different filter and does not have gen med or pulmonary subservice line breakdown
    so we are safe to exclude it from the logic of this function.  We will have to handle critical access in a different way.
    
    Takes a selenium driver variable as input and the measure name.
    returns the driver variable.
    
    '''
    time.sleep(0.3)
    if 'Sub-Service Line Not'.upper().strip() in browser_var.page_source.upper().strip() and '(GENERAL MEDICINE- COVID-19, GENERAL MEDICINE- GASTROENTEROLOGY)'.upper().strip() in browser_var.page_source.upper().strip():
    
        #only do this if the measure is Gen Med related.

        if measure_name_var in ['DCOST O/E - Medicine General','LOS O/E - Medicine General','Mortality O/E - Medicine General']:

            #step one.  click the edit image that is linked to subservice line restrictions.
            
            browser_var, subservice_line_edit = implicitly_wait_select_click_by_elem_type(browser_var, elem_obj = "//img[@title='Edit' and contains(@onclick, '12|UHCSubServiceLine')]", elem_type = By.XPATH)
            
            
            time.sleep(0.3)
            #step two.  select the master list of all subservice lines and scroll through it.  Unclick covid-19 gen med
            # and click the gen med - gastro.

            #master_list = WebDriverWait(browser_var, 3).until(find_adv_rest_master_list)
            
            browser_var, master_list = implicitly_wait_select_by_elem_type(browser_var, elem_obj = "lstMasterList", elem_type = By.ID)

            option_list1 = while_loop_handler_function(master_list.find_elements_by_tag_name('option'))

            pre_selected_options = [x.text for ind, x in enumerate(option_list1) if (ind==0) or (x.get_attribute("selected") == 'true')]

            #First go through and click the options that are in your period dictionary but not pre-selected.
            #We do this because when you click on the menu, it automatically clicks 
            #on the top option and does some funky stuff.  So, you need to first go through
            #and click on everything not pre-selected and not the first option. 
            #Then go back and click again or "unclick" the options that were pre-selected but not what you wanted.

            #In this case, we only want 'General Medicine- Gastroenterology'.  So, we are going to check it against the pre-selected
            #items and make sure it is selected.

            for optiona in option_list1:
                # Is the option from the drop-down at this point of iteration in our user-defined set of quarters? ALSO, is it NOT one of the following values?
                if optiona.text in ['General Medicine- Gastroenterology'] and optiona.text not in pre_selected_options:

                    attempt_option_click = 1
                    while (attempt_option_click != 0):
                        try:
                            optiona.click()  # select() in earlier versions of webdriver
                            attempt_option_click = 0
                        except StaleElementReferenceException:
                            if attempt_option_click == 3:
                                raise
                            attempt_option_click += 1
                            time.sleep(0.5)

            #now find all the menu options selected after the first round.
            selected_after_first_round = [x.text for ind, x in enumerate(option_list1) if (x.get_attribute("selected") == 'true')]

            #Unselected/Unclick by clicking again the items that are not in your period dictionary but are still
            #selected due to the preselection defaults.

            for optionb in option_list1:
                if optionb.text in selected_after_first_round and optionb.text not in ['General Medicine- Gastroenterology']:
                    attempt_option_click = 1
                    while (attempt_option_click != 0):
                        try:
                            #print('clicking on option.')
                            optionb.click()  # select() in earlier versions of webdriver
                            attempt_option_click = 0
                            #print('clicked on option.',optionb.text)
                        except StaleElementReferenceException:
                            if attempt_option_click == 3:
                                raise
                            attempt_option_click += 1
                            time.sleep(0.5)
            #click OK.
            #adv_rest_ok_bttn = WebDriverWait(browser_var, 3).until(find_adv_rest_save_bttn)
            #adv_rest_ok_bttn.click()
            
            browser_var, adv_rest_ok_bttn = implicitly_wait_select_click_by_elem_type(browser_var, elem_obj = "ctl00_ContentPlaceHolder1_ucctl11_btnSave", elem_type = By.ID)
            
    else:
        return(browser_var)
    
    return(browser_var)


################################################################################################

def remove_covid_19_pulmonary_subservice_rest(browser_var,measure_name_var):
    '''
    #this function is used to remove the covid 19 subservice line restriction on
    #only the pulmonary/critical care templates (but not readmission or edac for some reason...Just mort,los,dcost).

    #Critical access uses a different filter and does not have gen med or pulmonary subservice line breakdown
    #so we are safe to exclude it from the logic of this function.  We will have to handle critical access in a different way.
    
    #only do this if the measure is Pulmonary/Critical Care related.
    
    takes a selenium driver variable and measure name variable as input.
    Checks for existence of Pulmonary/ Critical Care -COVID-19 restriction.  If exists, deletes it.  If not, returns.
    returns selenium driver variable.
    '''
    
    if 'Sub-Service Line Not'.upper().strip() in browser_var.page_source.upper().strip() and 'Pulmonary/ Critical Care -COVID-19'.upper().strip() in browser_var.page_source.upper().strip():
    
        if measure_name_var in ['DCOST O/E - Pulmonary/Critical Care','LOS O/E - Pulmonary/Critical Care','Mortality O/E - Pulmonary/Critical Care']:

            #step one.  click the delete image that is linked to subservice line restrictions.
            browser_var, subservice_rest_delete = implicitly_wait_select_click_by_elem_type(browser_var, elem_obj = "//img[@title='Delete' and contains(@onclick, '12|UHCSubServiceLine')]", elem_type = By.XPATH)
            
        
    else:
        return(browser_var)
    
    return(browser_var)  

################################################################################################
    
def find_delete_discharge_month(driver):
    '''
    The next two functions are used to remove the discharge month restrictions on all the 
    EDAC and READM report templates.  Vizient added all these discharge month restrictions
    to force a certain time period on the report.  Obviously, if you want to pull numbers
    outside of that time period, the results will be blank.  So...we have to remove them from
    every edac and readm report template.  For some reason, this was only done to edac and readm.
    find and select all 'Delete' images that contain the javascript "DischargeMonth'
    
    '''
    #find and click the delete image for discharge month
    delete_discharge_month = driver.find_elements_by_xpath("//img[@title='Delete' and contains(@onclick, 'DischargeMonth')]")
    if delete_discharge_month:
        return(delete_discharge_month)
    else:
        return([])  
        
        
        
################################################################################################


def remove_covid_19_readm_edac_revisit_crit_acc_discharge_month(browser_var,measure_name_var,cohort_nm):
    
    '''
    if the measure name is edac, readm or revisits, then check whether there are any 
    restrictions for 'DischargeMonth.'  If you get any results, click the first element.
    Keep doing this until there aren't anymore on the DOM.
    
    Critical Access templates use a combination of discharge month restriction
    or subservice line or both.
    
    '''
    
    edac_readm_revisits = list_of_total_revisits_measures + ['OP Procedure Revisits - Urological',\
                           'OP Procedure Revisits - Colonoscopy',\
                            'OP Procedure Revisits - Biliary',\
                            'OP Procedure Revisits - Arthroscopy',\
                            'Urinary Procedures Revisits within 7-days',\
                            'Colonscopy Revisits within 7-days',\
                           'Arthroscopy Revisits within 7-days']
    
    
    #only do this if the measure is readm or edac
    if cohort_nm in ['Critical Access & Small Community'] or measure_name_var in edac_readm_revisits:
        
        time.sleep(0.3)
        #find any advanced restrictions associated with discharge month
        try:
            discharge_month = find_delete_discharge_month(browser_var)
        except:
            discharge_month = []
            return(browser_var)
        #Since the html ids and css change when things are added/removed to the DOM, 
        #we have to use a while loop and continuously check/delete while there are still
        #discharge month advanced restrictions.
        while len(discharge_month) > 0:
            try:
                discharge_month = find_delete_discharge_month(browser_var)
                discharge_month[0].click()
                time.sleep(0.3)
            except:
                discharge_month = []
        time.sleep(0.3)
        return(browser_var)
    else:
        time.sleep(0.3)
        return(browser_var)

################################################################################################

def find_delete_subservice_line_img2(driver):
    '''
    find and click the edit button assigned to SubServiceLine
    '''
    delete_subservice_line_img = driver.find_elements_by_xpath("//img[@title='Delete' and contains(@onclick, 'UHCSubServiceLine')]")
    if delete_subservice_line_img:
        return(delete_subservice_line_img)
    else:
        return([])   
        
        
################################################################################################
        
def remove_covid_19_crit_access_subservice_line(browser_var,cohort_nm):
    
    '''
    takes a selenium driver variable and cohort name string as input.
    if cohort is critical access, iteratively remove all sub-service line
    advanced restrictions.
    return the driver variable
    
    '''
    
    #only do this if cohort is critical access
    if cohort_nm in ['Critical Access & Small Community']:
        time.sleep(0.3)
        #find any advanced restrictions associated with discharge month
        try:
            subservice_line = find_delete_subservice_line_img2(browser_var)
        except:
            subservice_line = []
            return(browser_var)
        #Since the html ids and css change when things are added/removed to the DOM, 
        #we have to use a while loop and continuously check/delete while there are still
        #subservice line advanced restrictions.
        while len(subservice_line) > 0:
            try:
                subservice_line = find_delete_subservice_line_img2(browser_var)
                subservice_line[0].click()
                time.sleep(0.3)
            except:
                subservice_line = []
        time.sleep(0.3)
        return(browser_var)
    else:
        time.sleep(0.3)
        return(browser_var)
        

################################################################################################


def make_fy21_covid_changes(browser_obj_var, measure_name_var1, cohort_nm):
    
    '''
    takes a selenium driver variable, measure name variable, measure name variable (critical access) as input
    function puts all fy21 covid update functions together.
    returns the selenium driver variable
    '''
    #check for 'Edit Custom List' text.  If exists, remove the custom list.
    browser_obj_var = remove_custom_covid_list_from_template(browser_obj_var,measure_name_var1)
    #make sure all group by dropdown menu selections are back to pre-FY21 defaults.
    browser_obj_var = update_group_by_select_to_default(browser_obj_var,measure_name_var1)
    
    #This section checks whether the measure is a PSI.
    #If the measure is a PSI, then it will remove the first advanced restriction
    #on the report template.  This restriction is always the Covid-19 sub service line filter.
    #if the measure is not a psi, then this will do nothing.
    if measure_name_var1 in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
        
        browser_obj_var, first_psi_covid_adv_restriction = implicitly_wait_select_click_by_elem_type(browser_obj_var, elem_obj = '1_imgDelete', elem_type = By.ID)
    
    else:
        pass
    
    #check for 'Gen Med COVID-19' advanced restriction text.  If exists, remove this restriction.
    browser_obj_var = remove_covid_19_gen_med_subservice_rest(browser_obj_var,measure_name_var1)
    
    #check for "Pulmonary/ Critical Care -COVID-19" advanced restriction text.  If exists, remove this restriction.
    browser_obj_var = remove_covid_19_pulmonary_subservice_rest(browser_obj_var,measure_name_var1)
    
    #check for discharge-month-related advanced restrictions and iteratively remove them until all gone.
    #readm/edac, revisits measures or critical access cohort only
    browser_obj_var = remove_covid_19_readm_edac_revisit_crit_acc_discharge_month(browser_obj_var,measure_name_var1,cohort_nm)
    
    #remove all subservice line advanced restrictions for crtical access only.
    browser_obj_var = remove_covid_19_crit_access_subservice_line(browser_obj_var,cohort_nm)
    
    
    return(browser_obj_var)

################################################################################################
    
def choose_adjustment_model(browser_var2, link_dict, key):
    '''
    function takes a selenium driver variable, dictionary, iterator index as input.
    finds and clicks the appropriate risk model based on the dictionary value.
    returns the selenium driver variable.
    
    '''

    if link_dict[key][5].replace(' ','').upper() == '2022RISKMODEL(AMC)':
        
        browser_var2, adjustment_model_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radModifiedMSDRG"]', elem_type = By.XPATH)
       
    elif  link_dict[key][5].replace(' ','').upper() == '2022RISKMODEL(COMMUNITY)':
        
        browser_var2, adjustment_model_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radModifiedMSDRG_Commu"]', elem_type = By.XPATH)

    elif link_dict[key][5].replace(' ','').upper() == '2021RISKMODEL(AMC)':
        
        browser_var2, adjustment_model_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radMSDRG"]', elem_type = By.XPATH)
       
    elif  link_dict[key][5].replace(' ','').upper() == '2021RISKMODEL(COMMUNITY)':
        
        browser_var2, adjustment_model_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radMSDRG_Commu"]', elem_type = By.XPATH)

    elif link_dict[key][5].replace(' ','').upper() == '2020RISKMODEL(AMC)':
        
        browser_var2, adjustment_model_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radMSDRG"]', elem_type = By.XPATH)
       
    elif  link_dict[key][5].replace(' ','').upper() == '2020RISKMODEL(COMMUNITY)':
        
        browser_var2, adjustment_model_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radMSDRG_Commu"]', elem_type = By.XPATH)

    return (browser_var2)


################################################################################################
    
def choose_ahrq_version(browser_var2, link_dict, key):
    '''
    function takes a selenium driver variable, dictionary, iterator index as input.
    finds and clicks the appropriate ahrq version based on the dictionary value.
    returns the selenium driver variable.
    '''
    try:

        if link_dict[key][6] == 'V2022':
            # Click the Risk Adjustment Model radio button
            # MBecker_20211117: 2021 was added as the current so had to update 2020 as a previous
            #browser_var2, ahrq_version_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radAHRQCurrent"]', elem_type = By.XPATH)
            browser_var2, ahrq_version_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = "//*[@id='ctl00_ContentPlaceHolder1_radAHRQCurrent']", elem_type = By.XPATH)


        elif link_dict[key][6] == 'V2021':
            # Click the Risk Adjustment Model radio button
            # MBecker_20211117: 2021 was added as the current so had to update 2020 as a previous
            #browser_var2, ahrq_version_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radAHRQCurrent"]', elem_type = By.XPATH)
            browser_var2, ahrq_version_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = "//*[@id='ctl00_ContentPlaceHolder1_radAHRQPrevious']", elem_type = By.XPATH)
        
        elif link_dict[key][6] == 'V2020':
            # Click the Risk Adjustment Model radio button
            # MBecker_20211117: 2021 was added as the current so had to update 2020 as a previous
            #browser_var2, ahrq_version_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = '//*[@id="ctl00_ContentPlaceHolder1_radAHRQCurrent"]', elem_type = By.XPATH)
            browser_var2, ahrq_version_btn = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = "//*[@id='ctl00_ContentPlaceHolder1_radAHRQPrevious1']", elem_type = By.XPATH)
        
        
    except:
        print('Something is wrong with the AHRQ Version button.')
    return (browser_var2)


################################################################################################
    
def update_group_by_select(browser_var2,period_dict,period_type):
    
    '''
    function takes a selenium driver variable, period dictionary and period type as input.
    Function to select multiple group by
    EDAC and Readmission measure templates can only be filter by quarter so we need to group by discharge month
    in order to get monthly values.
    
    '''
    if period_type == 'QUARTER':
        browser_var2, multgroupbydiv = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = "ctl00_ContentPlaceHolder1_GroupByOutcomesWithMult", elem_type = By.ID)
        Select(browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_cmdMeasuresByWithMult']")).select_by_visible_text('Discharge Month 1st Admit')
        return(browser_var2)
    else:
        #update begin year
        return(browser_var2)
        
        
        
################################################################################################
        
def choose_adjustment_ahrq_update_groupby_scroll(browser_obj, hyper_dict, index_i, period_dict,idex_p):
    
    '''
    wrapper function takes a selenium driver variable, dictionary, index, period dictionary and period index as input
    runs 3 main functions to choose the risk model, choose the ahrq version, update the groupby menu for read/edac measures
    and scrolls down a little.
    returns the selenium driver variable.
    '''
    
    # Click Risk Adjustment Model button
    browser_obj = choose_adjustment_model(browser_obj, hyper_dict, index_i)
    
    # Click AHRQ Version button
    browser_obj = choose_ahrq_version(browser_obj, hyper_dict, index_i)
    
    # Update Multiple Group By Drop down
    browser_obj = update_group_by_select(browser_obj, period_dict, idex_p)
    
    ## scroll down
    browser_obj, div_element2 = implicitly_wait_select_by_elem_type(browser_obj, elem_obj = "//div[@id='ctl00_ContentPlaceHolder1_PanelContent3']", elem_type = By.XPATH)
                
    browser_obj.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element2)

    return(browser_obj)
    
################################################################################################
    
    
def update_time_period_select(browser_var2, period_dict, period_type):
    time.sleep(0.2)
    if period_type == 'QUARTER':
        el = while_loop_handler_function(
            browser_var2.find_element_by_id('ctl00_ContentPlaceHolder1_cmdTimePeriodQuarters'))

        # Select the element. This element is now a variable. The variable is in the drop-down?
        
        option_list1 = while_loop_handler_function(el.find_elements_by_tag_name('option'))

        # UL002
        # dynamically create a list of the pre-selected options in the Quarters menu because each calculator year
        # changes the defaults.  This will make the code more robust rather than hard coding quarters.
        pre_selected_options = [x.text for ind, x in enumerate(option_list1) if
                                (ind == 0) or (x.get_attribute("selected") == 'true')]
        # First go through and click the options that are in your period dictionary but not pre-selected.
        # We do this because when you click on the menu, it automatically clicks
        # on the top option and does some funky stuff.  So, you need to first go through
        # and click on everything not pre-selected and not the first option.
        # Then go back and click again or "unclick" the options that were pre-selected but not what you wanted.
        for optiona in option_list1:
            # Is the option from the drop-down at this point of iteration in our user-defined set of quarters? ALSO, is it NOT one of the following values?
            if optiona.text in period_dict['QUARTER'] and optiona.text not in pre_selected_options:

                attempt_option_click = 1
                while (attempt_option_click != 0):
                    try:
                        # print('clicking on option.')
                        optiona.click()  # select() in earlier versions of webdriver
                        attempt_option_click = 0
                        time.sleep(0.2)
                        #print('clicked on option.',optiona.text)
                        
                    except StaleElementReferenceException:
                        if attempt_option_click == 3:
                            raise
                        attempt_option_click += 1
                        time.sleep(0.2)
        # now find all the menu options selected after the first round.
        selected_after_first_round = [x.text for ind, x in enumerate(option_list1) if
                                      (x.get_attribute("selected") == 'true')]

        # Unselected/Unclick by clicking again the items that are not in your period dictionary but are still
        # selected due to the preselection defaults.

        for optionb in option_list1:
            if optionb.text in selected_after_first_round and optionb.text not in period_dict['QUARTER']:
                attempt_option_click = 1
                while (attempt_option_click != 0):
                    try:
                        optionb.click()  # select() in earlier versions of webdriver
                        attempt_option_click = 0
                        #print('clicked on option.', optionb.text)
                        time.sleep(0.2)
                    except StaleElementReferenceException:
                        if attempt_option_click == 3:
                            raise
                        attempt_option_click += 1
                        time.sleep(0.2)
    else:
        time.sleep(0.2)
        #click the "from year" menu
        browser_var2, fromyeardropdown = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = 'FromYearDiv', elem_type = By.ID)
        time.sleep(0.2)
        Select(
            browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_fromYear']")).select_by_visible_text(
            period_dict[period_type][0][1])
        time.sleep(0.2)
        #click the "from month" menu
        browser_var2, frommonthdropdown = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = 'FromMonthDiv', elem_type = By.ID)
        time.sleep(0.2)
        Select(browser_var2.find_element_by_xpath(
            "//*[@id='ctl00_ContentPlaceHolder1_fromMonth']")).select_by_visible_text(period_dict[period_type][0][0])
        # update end year
        time.sleep(0.2)
        # Select To Month before To Year in order avoid Vizient's popup window...
        #click the "to month" menu
        browser_var2, tomonthdropdown = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = 'ToMonthDiv', elem_type = By.ID)
        time.sleep(0.2)
        Select(
            browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toMonth']")).select_by_visible_text(
            period_dict[period_type][1][0])
        time.sleep(0.2)
        #Update To Year menu
        browser_var2, toyeardropdown = implicitly_wait_select_click_by_elem_type(browser_var2, elem_obj = 'ToYearDiv', elem_type = By.ID)
        time.sleep(0.2)
        Select(
            browser_var2.find_element_by_xpath("//*[@id='ctl00_ContentPlaceHolder1_toYear']")).select_by_visible_text(
            period_dict[period_type][1][1])
        time.sleep(0.2)



################################################################################################

def set_report_time_period(browser_obj, measure_name_var,period_dict,period_index):
    
    '''
    Function takes a selenium driver variable, a measure name string, time period dictionary, period type index variable as input.
    
    First clicks either the From Year: radio button or the Quarters radio button depending on report type.
    
    Then updates the either the dropdown menu or quarters select menu.
    
    '''
    
    if measure_name_var in list_of_total_revisits_measures:
        
            print('PERIOD DICT PASS TO SET_REPORT FUNCTION:',period_dict)
            
            try:
                
                browser_obj, quarters_period_radio_btn = implicitly_wait_select_click_by_elem_type(browser_obj, elem_obj = "ctl00_ContentPlaceHolder1_cmdTimePeriodQtr", elem_type = By.ID)
                time.sleep(1)
                update_time_period_select(browser_obj, period_dict, period_index)
                return(browser_obj)
            except:
                print('Cannot click on Quarters radio button.')
                update_time_period_select(browser_obj, period_dict, period_index)
                return(browser_obj)
    else:
        
        try:
            
            browser_obj, time_period_radio_btn = implicitly_wait_select_click_by_elem_type(browser_obj, elem_obj = "ctl00_ContentPlaceHolder1_cmdFromYear", elem_type = By.ID)
            update_time_period_select(browser_obj, period_dict, period_index)
            return(browser_obj)
            
            
        except:
            
            try:
                
                browser_obj, time_period_radio_btn = implicitly_wait_select_click_by_elem_type(browser_obj, elem_obj = "ctl00_ContentPlaceHolder1_fromYear", elem_type = By.ID)
                update_time_period_select(browser_obj, period_dict, period_index)
                return(browser_obj)
                
            except:
                
                print('Cannot click on To/From radio button.')
                return(browser_obj)
            

################################################################################################

def scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(browser_helper_obj):
    
                '''
                Function takes a selenium driver variable as input.
                scrolls down to middle of the page to ensure elements are loaded.
                Sets the focus hospital menu to "Northwestern Memorial Hospital"
                Clicks the "all hospitals in database radio button"
                scrolls down
                returns the selenium driver variable
                '''
        
                #scroll down
                browser_helper_obj, div_element3 = implicitly_wait_select_by_elem_type(browser_helper_obj, elem_obj = "//div[@id='ctl00_ContentPlaceHolder1_PanelContent4']", elem_type = By.XPATH)
                
                browser_helper_obj.execute_script(
                    "return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element3)

                time.sleep(1.5)
                
                # Set focus hospital to NMH
                #browser_helper_obj, focus_hosp = implicitly_wait_select_click_by_elem_type(browser_helper_obj, elem_obj = "//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']", elem_type = By.XPATH)
                
                try:
                    #browser_helper_obj.find_element(By.XPATH, "//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']/option[263]")
                    nmh_hco_id = Select(browser_helper_obj.find_element(By.ID, "ctl00_ContentPlaceHolder1_cmdFocusHCO"))
                    nmh_hco_id.select_by_value("140281")  #nmh
                    #Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')

                    
                except:
                    
                    #browser_helper_obj, focus_hosp = implicitly_wait_select_click_by_elem_type(browser_helper_obj, elem_obj = "//*[@id='ctl00_ContentPlaceHolder1_cmdFocusHCO']", elem_type = By.XPATH)
                    #Select(focus_hosp).select_by_visible_text('NORTHWESTERN_MEMORIAL 140281')

                    nmh_hco_id = Select(browser_helper_obj.find_element(By.ID, "ctl00_ContentPlaceHolder1_cmdFocusHCO"))
                    nmh_hco_id.select_by_value("140281")  #nmh
                
                
                browser_helper_obj, div_element4 = implicitly_wait_select_by_elem_type(browser_helper_obj, elem_obj = "//div[@id='ctl00_ContentPlaceHolder1_PanelContent5']", elem_type = By.XPATH)
                
                browser_helper_obj.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element4)

                
                #select 'all hospitals in database'
                browser_helper_obj, focus_hosp = implicitly_wait_select_click_by_elem_type(browser_helper_obj, elem_obj = "ctl00_ContentPlaceHolder1_cmdAllAvailHosp", elem_type = By.ID)
                

                # scroll down
                
                browser_helper_obj, div_element5 = implicitly_wait_select_by_elem_type(browser_helper_obj, elem_obj = "//div[@id='divRiskAdjustment']", elem_type = By.XPATH)
                
                browser_helper_obj.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)
                
                time.sleep(1)
                return(browser_helper_obj)
                
                
################################################################################################
                
def add_covid_advanced_restriction(driver_var,meas_type = 'OTHER'):
    
    '''
    Function used to add COVID Diagnosis Code NOT (U071) advanced restriction.  This is used to 'remove' COVID
    patients where possible.  Takes a selenium driver variable as input and measure type variable as input.
    returns the selenium driver variable.
    
    '''
    
    if meas_type == 'THK':
        
        return(driver_var)
    else:
        print('other...')
        #click the advanced restrictions menu
        time.sleep(0.5)
        adv_rest = WebDriverWait(driver_var, 10).until(find_adv_restrictions_add)
        adv_rest.click()
        time.sleep(3)


        # click the 'Diagnosis/Procedure option'
        #adv_rest_menu_id = WebDriverWait(driver_var, 10).until(find_adv_restrictions_menu_id)
        #print(adv_rest_menu_id)
        time.sleep(0.5)
        diag_proc_option = WebDriverWait(driver_var, 10).until(find_diag_proc_option)
        time.sleep(1)
        diag_proc_option.click()

        '''
        #time.sleep(0.5)
        rad_span_list = adv_rest_menu_id.find_elements_by_xpath('.//span[@class = "rmText rmExpandRight"]')
        print(rad_span_list)
        #rad_span_list = find_rad_span_list(driver_var,adv_rest_menu_id)
        #print(rad_span_list)
        #loop over spans under the advanced restriction menu and only click the one that has text 'Diagnosis/Procedure'
        for i in rad_span_list:
            if i.text == 'Diagnosis / Procedure':
                i.click()
        #time.sleep(0.5)
        #click the 'Any Diagnosis' option
        '''
        time.sleep(0.5)
        any_diag_option = WebDriverWait(driver_var, 10).until(find_any_diag_option) 
        time.sleep(1)
        any_diag_option.click()
        time.sleep(1)
        #a popup window will appear.  Switch control to the new window.

        browser_driver1 = driver_var.window_handles[0]
        browser_driver2 = driver_var.window_handles[1]

        driver_var.switch_to.window(browser_driver2)

        #click the 'Exclude Only' dropdown menu option
        exclude_only_option = WebDriverWait(driver_var, 10).until(find_exclude_only)
        time.sleep(0.5)
        exclude_only_option.click()

        #Type the U071 ICD-10 code into the search bar to find the Covid ICD-10 code
        search_bar = WebDriverWait(driver_var, 10).until(find_search_bar)
        time.sleep(0.5)
        #Type Covid ICD10 Code
        search_bar.send_keys("U071")
        time.sleep(1)
        #click 'Search'
        search_bttn = WebDriverWait(driver_var, 10).until(find_search_bttn_adv_rest)
        time.sleep(0.5)
        search_bttn.click()
        time.sleep(0.5)
        #select the U071 option
        all_box_bttn = WebDriverWait(driver_var, 10).until(find_all_box_bttn)
        all_box_bttn.click()
        time.sleep(1)
        #click 'Add Selections'
        add_selections_bttn = WebDriverWait(driver_var, 10).until(find_add_selections_bttn)
        time.sleep(0.5)
        add_selections_bttn.click()
        time.sleep(0.5)
        #click 'Send to CDB'
        send_to_cdb_bttn = WebDriverWait(driver_var, 10).until(find_send_to_cdb_bttn)
        time.sleep(0.5)
        send_to_cdb_bttn.click()

        #return driver control back to first window
        driver_var.switch_to.window(browser_driver1)
        
        if meas_type == 'READM':
            time.sleep(0.5)
            
            adv_rest_apply_index_menu = WebDriverWait(driver_var, 10).until(find_adv_rest_index_apply)
            
            
            for i, item in enumerate(adv_rest_apply_index_menu[0].find_elements_by_tag_name('option')):
                
                if item.text == 'Both':
                    item.click()
            return(driver_var)
        else:
            return(driver_var) 



################################################################################################


def click_excel_button_handle_popups(browser_obj_helper,download_dir,num_downloaded_files):
    
    '''
    function takes a selenium driver variable as input
    clicks the excel icon button to download the report
    clicks accept to close popups and close a second browser window that opens upon download
    returns the selenium driver variable.
    
    '''
    time.sleep(1)
    #click excel icon to download the data
    browser_obj_helper, generate_excel_btn = implicitly_wait_select_click_by_elem_type(browser_obj_helper, elem_obj = 'ctl00_ContentPlaceHolder1_imgExcel', elem_type = By.ID)
    
    #when you click download, a popup pops up.  click accept to close.
    try:
        WebDriverWait(browser_obj_helper, 10).until(EC.alert_is_present(),'Timed out waiting for PA creation ' +'confirmation popup to appear.')

        obj = WebDriverWait(browser_obj_helper, 10).until(find_alert_popup)

        obj.accept()
    except:
        pass
    
    #wait until the number of files in the downloads folder changes
    while len(os.listdir(download_dir)) <= num_downloaded_files:
        time.sleep(1)
        
    # wait for the pop up browser window to display so you can close it.
    while len(browser_obj_helper.window_handles) < 2:
        time.sleep(0.5)
        
    #when multiple browsers open up when the report is downloaded, switch control to the new window.  Close it.  Then 
    #switch back to the first browser window.
    window_before = browser_obj_helper.window_handles[0]
    window_after = browser_obj_helper.window_handles[1]
    # switch control to popup window and close it
    browser_obj_helper.switch_to.window(window_after)
    browser_obj_helper.close()
    # Switch control back to the original window.
    browser_obj_helper.switch_to.window(window_before)
    time.sleep(2)
    return(browser_obj_helper,download_dir,num_downloaded_files)

################################################################################################
    
def find_last_downloaded_file(dir):
    
    '''
    function takes a file directory as input and returns the file name of the most recently downloaded file.
    '''
    
    list_of_files = glob.glob(dir + '/*')
    latest_file = max(list_of_files, key=os.path.getctime)
    ##UL008
    while latest_file.endswith('.crdownload') == True:
        print('accidentally grabbed crdownload temp file.  Trying again...')
        time.sleep(1)
        list_of_files = glob.glob(dir + '/*')
        latest_file = max(list_of_files, key=os.path.getctime)
    print(latest_file)
    return(latest_file)  
    
################################################################################################

def rename_and_move_file(file, hospital_type, measure_name, period_type, new_file_dir):
    # os.chdir('C:/Data/Downloads')
    '''
    Function takes a file, renames it using the hospital cohort type and measure name.
    '''
    
    if hospital_type == 'Complex Care Medical Center':
        hospital_type2 = 'CCMC'
    elif hospital_type == 'Comprehensive Academic Medical Center':
        hospital_type2 = 'AMC'
    elif (hospital_type == 'Large Specialized Complex Care Medical Center') or (hospital_type == 'Large, Specialized Complex Care Medical Center'):
        hospital_type2 = 'LSCCMC'
    #UL002 add critical access to file naming options
    elif hospital_type == 'Critical Access & Small Community':
        hospital_type2 = 'CASC'
    else:
        hospital_type2 = 'COMM'

    new_hospital_name = (str(hospital_type2).replace(" ", "_")).replace("-", "_").upper()
    new_measure_name = (str(measure_name).replace(" ", "_")).replace("-", "_").upper()
    new_period_type = (str(period_type).replace(" ", "_")).replace("-", "_").upper()
    new_file_name = new_hospital_name + '_' + new_measure_name + '_' + new_period_type + '.xlsx'

    new_path = os.path.join(new_file_dir, str(hospital_type), str(new_measure_name), new_file_name)
    main_cohort_path = os.path.join(new_file_dir, str(hospital_type), str(new_measure_name))

    if not os.path.exists(main_cohort_path):
        os.makedirs(main_cohort_path)
    if os.path.exists(new_path):
        print('File already exists!')
        print(new_path)
        overwrite_decision = input("Do you really want to overwrite this file? Choose: 'YES' or 'NO'")

        if overwrite_decision.upper() == 'YES':
            shutil.move(file, new_path)
        else:
            print('Please handle the existing file or change the directory location.')
            exit()
    shutil.move(file, new_path) 
    
################################################################################################   
    
def update_template_files(hyperlink_loc, cohort, measure):
    
    '''
    Function to update the template excel files.  This will be used if the program crashes midway through
    and you need to pick up where you left off.  As reports are successfully downloaded, they will be removed
    from the template excel files.
    
    '''
    
    if cohort == 'Comprehensive Academic Medical Center':
        time.sleep(1)
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'nmh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('nmh_links.xlsx')
            dataframe_ob.to_excel(writer,'Sheet1', index=False)
            writer.save()
        except:
            print('Issue finding nmh_links.xlsx.')
            pass
    elif cohort == 'Large Specialized Complex Care Medical Center':
        time.sleep(1)
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'cdh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('cdh_links.xlsx')
            dataframe_ob.to_excel(writer,'Sheet1', index=False)
            writer.save()
        except:
            print('Issue finding cdh_links.xlsx.')
            pass
    elif cohort == 'Complex Care Medical Center':
        time.sleep(1)
        try:
            #UL007
            dataframe_ob3 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'lfh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob3 = dataframe_ob3[dataframe_ob3['Formal Name'] != measure]
            writer3 = pd.ExcelWriter('lfh_links.xlsx')
            dataframe_ob3.to_excel(writer3,'Sheet1', index=False)
            writer3.save()
            #UL007
            dataframe_ob1 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'dch_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob1 = dataframe_ob1[dataframe_ob1['Formal Name'] != measure]
            writer1 = pd.ExcelWriter('dch_links.xlsx')
            dataframe_ob1.to_excel(writer1,'Sheet1', index=False)
            writer1.save()
            #UL007
            dataframe_ob2 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'kish_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            dataframe_ob2 = dataframe_ob2[dataframe_ob2['Formal Name'] != measure]
            writer2 = pd.ExcelWriter('kish_links.xlsx')
            dataframe_ob2.to_excel(writer2,'Sheet1', index=False)
            writer2.save()
            
            #UL007
            dataframe_ob4 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'hh_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob4 = dataframe_ob4[dataframe_ob4['Formal Name'] != measure]
            writer4 = pd.ExcelWriter('hh_links.xlsx')
            dataframe_ob4.to_excel(writer4,'Sheet1', index=False)
            writer4.save()
            #UL007
            dataframe_ob5 = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc),'mch_links.xlsx'),sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob5 = dataframe_ob5[dataframe_ob5['Formal Name'] != measure]
            writer5 = pd.ExcelWriter('mch_links.xlsx')
            dataframe_ob5.to_excel(writer5,'Sheet1', index=False)
            writer5.save()
        except:
            print('Issue finding CCMC link files for update.')
            pass
    elif cohort == 'Community Medical Center':
        time.sleep(1)
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc), 'comm_links.xlsx'), sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('comm_links.xlsx')
            dataframe_ob.to_excel(writer, 'Sheet1', index=False)
            writer.save()
        except:
            print('Issue finding comm_links.xlsx.')
            pass
    #UL002 adding new critical access cohort to file updater
    elif cohort == 'Critical Access & Small Community':
        time.sleep(1)
        try:
            #UL007
            dataframe_ob = pd.DataFrame(pd.read_excel(os.path.join(os.path.abspath(hyperlink_loc), 'vwh_links.xlsx'), sheet_name="Sheet1",engine='openpyxl'))
            time.sleep(0.5)
            dataframe_ob = dataframe_ob[dataframe_ob['Formal Name'] != measure]
            writer = pd.ExcelWriter('vwh_links.xlsx')
            dataframe_ob.to_excel(writer, 'Sheet1', index=False)
            writer.save()
        except:
            print('Issue finding vwh_links.xlsx.')
            pass    
    
    
################################################################################################ 
    
      
def loop_template_download_fy22(link_dict, period_dict1,period_dict2,period_dict3, driver_var, exclusion_list,file_dir,hyperlink_loc,remove_covid_pats = False):
    
    '''
    function takes our cohort/hyperlink/measure dictionary as input, takes our time period dictionaries as lookup,
    our selenium driver variable as input.
    
    Loops over the hyperlink dictionary and conditionally opens each hyperlink, clicks all the necessary input buttons,
    conditionally modifies reports, downloads the report, renames the report and moves the report to the correct
    folder within the output file location.
    
    '''
    
    
    #initialize report counter so we can report out how many reports were downloaded at the end.
    report_counter = 0
    
    #find the path to the downloads folder
    download_folder_dir = grab_download_folder_dir()
                    
    for i in link_dict.keys():
        if link_dict[i][4] == 'THK Complication':
            
            #THK measure only
            
            for p2 in period_dict2.keys():
                
                print('MEASURE:',link_dict[i][4])
                
                #check number of files in the downloads folder, open the cdb report template, scroll to the top
                driver_var,num_files = check_num_downloaded_files_open_template_max_screen_and_scroll(driver_var,download_folder_dir,link_dict[i][0])
                
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #driver_var = make_fy21_covid_changes(driver_var,link_dict[i][4],link_dict[i][3])
                
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                #update the risk adjustment model, ahrq version, groupby selection, scroll down a little
                driver_var = choose_adjustment_ahrq_update_groupby_scroll(driver_var, link_dict, i, period_dict2, p2)
                
                #set the time period
                set_report_time_period(driver_var, link_dict[i][4],period_dict2, p2)
                
                #scroll, select focus hospital, select 'all hospitals in database', scroll
                driver_var = scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(driver_var)
                
                #remove covid patients by adding an advanced restriction.
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var,'THK')
                    
                driver_var,download_folder_dir,num_files  = click_excel_button_handle_popups(driver_var,download_folder_dir,num_files)
                
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184797/Downloads')
                    
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p2, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(1)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(1)

        elif link_dict[i][4] in list_of_total_revisits_measures:
            
            # READM/EDAC measures
            
            for p3 in period_dict3.keys():
                
                print('MEASURE:',link_dict[i][4])
                
                #check number of files in the downloads folder, open the cdb report template, scroll to the top
                driver_var,num_files = check_num_downloaded_files_open_template_max_screen_and_scroll(driver_var,download_folder_dir,link_dict[i][0])
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #driver_var = make_fy21_covid_changes(driver_var,link_dict[i][4],link_dict[i][3])
                
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                #update the risk adjustment model, ahrq version, groupby selection, scroll down a little
                driver_var = choose_adjustment_ahrq_update_groupby_scroll(driver_var, link_dict, i, period_dict3, p3)
                
                #set the time period
                set_report_time_period(driver_var, link_dict[i][4],period_dict3, p3)
                
                #scroll, select focus hospital, select 'all hospitals in database', scroll
                driver_var = scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(driver_var)
                
                #remove covid patients by adding an advanced restriction.
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var,'READM')
                    
                driver_var,download_folder_dir,num_files  = click_excel_button_handle_popups(driver_var,download_folder_dir,num_files)
                
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184797/Downloads')
                    
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p3, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(2)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(2)
                
                
        # leave a filter in the code for now until Vizient resolves report radio button issue...
        #print(link_dict[i][4])
        elif link_dict[i][4] not in list_of_total_revisits_measures and link_dict[i][4] in ['PSI-03 O/E','PSI-11 O/E','PSI-09 O/E','PSI-06 O/E','PSI-13 O/E']:
            
            #PSI MEASURES
            
            for p1 in period_dict1.keys():
                
                print('MEASURE:',link_dict[i][4])
                
                #check number of files in the downloads folder, open the cdb report template, scroll to the top
                driver_var, num_files = check_num_downloaded_files_open_template_max_screen_and_scroll(driver_var,download_folder_dir,link_dict[i][0])
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #driver_var = make_fy21_covid_changes(driver_var,link_dict[i][4],link_dict[i][3])
                
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                #update the risk adjustment model, ahrq version, groupby selection, scroll down a little
                driver_var = choose_adjustment_ahrq_update_groupby_scroll(driver_var, link_dict, i, period_dict1, p1)
                
                #set the time period
                set_report_time_period(driver_var, link_dict[i][4],period_dict1, p1)
                
                #scroll, select focus hospital, select 'all hospitals in database', scroll
                driver_var = scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(driver_var)
                
                # scroll down
                driver_var, adv_rest_div_element = implicitly_wait_select_by_elem_type(driver_var, elem_obj = "//div[@id='tblAdvRestrictionsDiv']", elem_type = By.XPATH)
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",adv_rest_div_element)

                #click first advanced restriction
                driver_var, generate_delete_btn = implicitly_wait_select_click_by_elem_type(driver_var, elem_obj = "1_imgDelete", elem_type = By.ID)
                
                #scroll
                driver_var, div_element5 = implicitly_wait_select_by_elem_type(driver_var, elem_obj = "//div[@id='divRiskAdjustment']", elem_type = By.XPATH)
                driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)

                #remove covid patients by adding an advanced restriction.
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var)
                    
                driver_var,download_folder_dir,num_files  = click_excel_button_handle_popups(driver_var,download_folder_dir,num_files)
                
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184797/Downloads')
                    
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(1)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(1)
                
        #UL003
        # Critical Access Measures:  Adverse Drug Events Rate & % Early Transfers Out
        # Require their reports run twice...
        # Run once with default settings to get numerator (Cases)
        # Run second time removing all Advanced Restrictions to get denominator (Cases)
        # Once reports are downloaded, then will need to take numerator/denominator to get true result.
        elif link_dict[i][4] in list_of_measures_to_run_twice_num_denom:
            for p1 in period_dict1.keys():
                #run twice.
                for item in ['NUM','DENOM']:
                    #first time run in default setting to get numerator
                    if item == 'NUM':
                        #if item is 'NUM' then run the ADE or % Early Transfer report in default mode.
                        # The 'Cases' column will give the numerator of the ratio.
                        
                        print('MEASURE:',link_dict[i][4])
                        print('NUMERATOR')
                        #check number of files in the downloads folder, open the cdb report template, scroll to the top
                        driver_var, num_files = check_num_downloaded_files_open_template_max_screen_and_scroll(driver_var,download_folder_dir,link_dict[i][0])
                        
                        #BEGIN FY21 COVID FILTER FUNCTIONS#
                        ###################################
                
                        #driver_var = make_fy21_covid_changes(driver_var,link_dict[i][4],link_dict[i][3])
                
                        #END FY21 COVID FILTER FUNCTIONS#
                        ##################################
                        
                        #update the risk adjustment model, ahrq version, groupby selection, scroll down a little
                        driver_var = choose_adjustment_ahrq_update_groupby_scroll(driver_var, link_dict, i, period_dict1, p1)
                        
                        #set the time period
                        set_report_time_period(driver_var, link_dict[i][4],period_dict1, p1)
                        
                        #scroll, select focus hospital, select 'all hospitals in database', scroll
                        driver_var = scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(driver_var)
                        
                        #remove covid patients by adding an advanced restriction.
                        if remove_covid_pats == True:
                            driver_var = add_covid_advanced_restriction(driver_var)
                            
                        driver_var,download_folder_dir,num_files  = click_excel_button_handle_popups(driver_var,download_folder_dir,num_files)
                        
                        time.sleep(2)
                        # find most recently-added file to Downloads folder and rename it.
                        try:
                            latest_file = find_last_downloaded_file('C:/Users/NM184797/Downloads')
                            
                        except:
                            latest_file = find_last_downloaded_file('H:/Downloads')
                        latest_file = os.path.abspath(latest_file)
        
                        time.sleep(0.500)
                        
                        #UL003
                        #if 'NUM' then rename the measure_name to ..._NUM.  For example, ADE_NUM.
                        #UL003
                        updated_measure_name = link_dict[i][7] + '_NUM'
                        #UL003
                        #rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                        rename_and_move_file(latest_file, link_dict[i][3], updated_measure_name, p1, file_dir)
                        time.sleep(0.500)
                        report_counter += 1
                        time.sleep(2)
                        update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                        time.sleep(2)
                        
                if item == 'DENOM':
                    
                    
                        #Adverse Drug Events only has one advanced restriction so only need
                        #delete one advanced restriction.
                        if link_dict[i][4] == 'Adverse Drug Events Rate':
                            
                            print('MEASURE:',link_dict[i][4])
                            print('DENOMINATOR')
                            #check number of files in the downloads folder, open the cdb report template, scroll to the top
                            driver_var, num_files = check_num_downloaded_files_open_template_max_screen_and_scroll(driver_var,download_folder_dir,link_dict[i][0])
                            
                            #BEGIN FY21 COVID FILTER FUNCTIONS#
                            ###################################
                
                            #driver_var = make_fy21_covid_changes(driver_var,link_dict[i][4],link_dict[i][3])
                
                            #END FY21 COVID FILTER FUNCTIONS#
                            ##################################
                        
                            #update the risk adjustment model, ahrq version, groupby selection, scroll down a little
                            driver_var = choose_adjustment_ahrq_update_groupby_scroll(driver_var, link_dict, i, period_dict1, p1)
                            
                            #set the time period
                            set_report_time_period(driver_var, link_dict[i][4],period_dict1, p1)
                            
                            #scroll, select focus hospital, select 'all hospitals in database', scroll
                            driver_var = scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(driver_var)
                            
                            
                            # Similar to PSI's, we need to remove the advanced restrictions
                            # For some reason, on the Critical Access calculator, 2 measures (ADE, % Early Transfers)
                            # the report template does not give the final correct calculations.
                            # According to Vizient, you actually have to run the report twice. 
                            # Adverse Drug Events Rate only has 1 advanced restriction so we only need to 
                            # Click on 1 delete button in the advanced restrictions section similar to the PSI measures.
                            
                            # scroll down
                            driver_var, adv_rest_div_element = implicitly_wait_select_by_elem_type(driver_var, elem_obj = "//div[@id='tblAdvRestrictionsDiv']", elem_type = By.XPATH)
                
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",adv_rest_div_element)

                            #click first advanced restriction
                            driver_var, generate_delete_btn = implicitly_wait_select_click_by_elem_type(driver_var, elem_obj = "1_imgDelete", elem_type = By.ID)
                
                            #scroll
                            driver_var, div_element5 = implicitly_wait_select_by_elem_type(driver_var, elem_obj = "//div[@id='divRiskAdjustment']", elem_type = By.XPATH)
                
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)

                            #remove covid patients by adding an advanced restriction.
                            if remove_covid_pats == True:
                                driver_var = add_covid_advanced_restriction(driver_var)
                                
                            driver_var,download_folder_dir,num_files  = click_excel_button_handle_popups(driver_var,download_folder_dir,num_files)
                            
                            time.sleep(2)
                            # find most recently-added file to Downloads folder and rename it.
                            try:
                                latest_file = find_last_downloaded_file('C:/Users/NM184797/Downloads')
                                
                            except:
                                latest_file = find_last_downloaded_file('H:/Downloads')
                            latest_file = os.path.abspath(latest_file)
            
                            time.sleep(0.500)
                            
                            #UL003
                            #if 'DENOM' then rename the measure_name to ..._NUM.  For example, ADE_NUM.
                            #UL003
                            updated_measure_name = link_dict[i][7] + '_DENOM'
                            #UL003
                            #rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                            rename_and_move_file(latest_file, link_dict[i][3], updated_measure_name, p1, file_dir)
                            time.sleep(0.500)
                            report_counter += 1
                            time.sleep(2)
                            update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                            time.sleep(2)
                            
                        #% Early Transfer Out has 4 advanced restrictions so we need to remove
                        #4 different advanced restrictions by clicking the delete button.
                        elif link_dict[i][4] == '% Early Transfers Out':
                            
                            print('MEASURE:',link_dict[i][4])
                            print('link var at the top',link_dict[i])
                            #check number of files in the downloads folder, open the cdb report template, scroll to the top
                            driver_var, num_files = check_num_downloaded_files_open_template_max_screen_and_scroll(driver_var,download_folder_dir,link_dict[i][0])
                            
                            #BEGIN FY21 COVID FILTER FUNCTIONS#
                            ###################################
                
                            #driver_var = make_fy21_covid_changes(driver_var,link_dict[i][4],link_dict[i][3])
                
                            #END FY21 COVID FILTER FUNCTIONS#
                            ##################################
                        
                            #update the risk adjustment model, ahrq version, groupby selection, scroll down a little
                            driver_var = choose_adjustment_ahrq_update_groupby_scroll(driver_var, link_dict, i, period_dict1, p1)
                            
                            #set the time period
                            set_report_time_period(driver_var, link_dict[i][4],period_dict1, p1)
                            
                            #scroll, select focus hospital, select 'all hospitals in database', scroll
                            driver_var = scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(driver_var)
                            
                            # scroll down
                            driver_var, adv_rest_div_element = implicitly_wait_select_by_elem_type(driver_var, elem_obj = "//div[@id='tblAdvRestrictionsDiv']", elem_type = By.XPATH)
                
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);",adv_rest_div_element)
                            print('before removing early transfer adv. restrictions')
                            print('link var in the middle',link_dict[i])
                            #click 4 advanced restrictions.  Start with 4th restriction, then 3rd, 2nd, 1st.
                            for rest in range(4,0,-1):
                                driver_var, generate_delete_btn = implicitly_wait_select_click_by_elem_type(driver_var, elem_obj = str(rest)+"_imgDelete", elem_type = By.ID)
                                time.sleep(0.3)
                                
                            #scroll
                            driver_var, div_element5 = implicitly_wait_select_by_elem_type(driver_var, elem_obj = "//div[@id='divRiskAdjustment']", elem_type = By.XPATH)
                
                            driver_var.execute_script("return arguments[0].scrollIntoView(0, document.documentElement.scrollHeight-10);", div_element5)

                            #remove covid patients by adding an advanced restriction.
                            if remove_covid_pats == True:
                                driver_var = add_covid_advanced_restriction(driver_var)
                                
                            print('before clicking excel button.  early transfers')
                            driver_var,download_folder_dir,num_files  = click_excel_button_handle_popups(driver_var,download_folder_dir,num_files)
                            
                            time.sleep(2)
                            # find most recently-added file to Downloads folder and rename it.
                            try:
                                latest_file = find_last_downloaded_file('C:/Users/NM184797/Downloads')
                                
                            except:
                                latest_file = find_last_downloaded_file('H:/Downloads')
                            latest_file = os.path.abspath(latest_file)
                            print('found latest file')
                            time.sleep(0.500)
                            print('link var dict: \n',link_dict[i])
                            #UL003
                            #if 'DENOM' then rename the measure_name to ..._DENOM.  For example, ADE_DENOM.
                            #UL003
                            updated_measure_name = link_dict[i][7] + '_DENOM'
                            #UL003
                            #rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                            rename_and_move_file(latest_file, link_dict[i][3], updated_measure_name, p1, file_dir)
                            time.sleep(0.500)
                            report_counter += 1
                            time.sleep(2)
                            update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                            time.sleep(2)
                           
        elif link_dict[i][4] not in list_of_total_revisits_measures:
            
            #All other report templates:  MORT, LOS, DCOST
            
            for p1 in period_dict1.keys():
                
                print('MEASURE:',link_dict[i][4])
                
                #check number of files in the downloads folder, open the cdb report template, scroll to the top
                driver_var, num_files = check_num_downloaded_files_open_template_max_screen_and_scroll(driver_var,download_folder_dir,link_dict[i][0])
                
                #BEGIN FY21 COVID FILTER FUNCTIONS#
                ###################################
                
                #driver_var = make_fy21_covid_changes(driver_var,link_dict[i][4],link_dict[i][3])
                
                #END FY21 COVID FILTER FUNCTIONS#
                ##################################
                
                #update the risk adjustment model, ahrq version, groupby selection, scroll down a little
                driver_var = choose_adjustment_ahrq_update_groupby_scroll(driver_var, link_dict, i, period_dict1, p1)
                
                #set the time period
                set_report_time_period(driver_var, link_dict[i][4],period_dict1, p1)
                
                #scroll, select focus hospital, select 'all hospitals in database', scroll
                driver_var = scroll_down_set_focus_hosp_scoll_select_all_hosp_scoll(driver_var)
                
                #remove covid patients by adding an advanced restriction.
                if remove_covid_pats == True:
                    driver_var = add_covid_advanced_restriction(driver_var)
                    
                driver_var,download_folder_dir,num_files  = click_excel_button_handle_popups(driver_var,download_folder_dir,num_files)
                
                time.sleep(2)
                # find most recently-added file to Downloads folder and rename it.
                try:
                    latest_file = find_last_downloaded_file('C:/Users/NM184797/Downloads')
                    
                except:
                    latest_file = find_last_downloaded_file('H:/Downloads')
                latest_file = os.path.abspath(latest_file)

                time.sleep(0.500)

                rename_and_move_file(latest_file, link_dict[i][3], link_dict[i][7], p1, file_dir)
                time.sleep(0.500)
                report_counter += 1
                time.sleep(2)
                update_template_files(hyperlink_loc, link_dict[i][3], link_dict[i][4])
                time.sleep(2)
                
    return (report_counter)
                
    
    
 ################################################################################################   
    
    

# define one function to put it all together.
#UL006  adding new parameter to control whether or not we remove covid patients.

#UL010  Updating this all functions within core_scraper_function for fy22
#       adding documentation, removing duplication.
def core_scraper_function(remove_covid = False):
    
    '''
    This is the main selenium web scraper function that opens Vizient, loops over a list of Q&A report template
    hyperlinks, downloads each report, sorts the output reports into correct folders in 230 Inpatient Quality Composite.
    
    # step 1: What period is this for?  Enter in the end date of for the Performance Close month.  Import
    #         the time period data and generate a helper dictionary from it.
    
    # step 2:  Import cohort data from the Vizient documentation file.
    
    # step 3:  Import the Vizient template hyperlink file and generate a helper dictionary from it.
    
    # step 4:  Create main folder structure to store the files in.
    
    # step 5:  Setup selenium driver and Login to Vizient main page
    
    '''
    
    #start out by deciding where we want to get out input commands from.
    #For FY22, I created a new function which allows us to write all the input 
    #time periods and filepaths into an excel file and read from that instead of typing many inputs strings.
    input_instructions_choice = '0'

    while input_instructions_choice not in ['1','2']:
        input_instructions_choice = input(
        "Type the number 1 if you would like to use the core_scraper_function_input_instructions file.  Type the number 2 if you would like to manually input time periods and filepaths using an input prompt.")
    
    if input_instructions_choice == '1':
        
        #if we choose option one, open the core_scraper_inputs.xlsx file and get all input data.
        #create a dictionary with it.
        core_input_dict = collect_core_scraper_inputs()
    
    else:
        core_input_dict = {}
    #core_scraper_function is a wrapper function for all
    #functions that login to vizient, build a cohort/time period dictionary, 
    #loop over all cohort reports and save them to a folder structure.
    
    
    # step 1: What period is this for?  Enter in the end date of for the Performance Close month.  Import
    #         the time period data and generate a helper dictionary from it.
    print('Choose time period for most measures.')
    #all "most measures" metrics aka 1-month lag measure begin/end datetimes.
    if input_instructions_choice == '1':
        period_helper_dict1 = build_period_lookup_dict(input_dict = core_input_dict, input_dict_key_list=['most_measures_begin_date','most_measures_end_datetime'],time_period_choice_num='2')
    else:
        period_helper_dict1 = build_period_lookup_dict()
        
    print('DICT 1:',period_helper_dict1)
        
    print('Choose time period for THK measure.')
    #THK and other 2-month lag metric begin/end datetimes
    if input_instructions_choice == '1':
        period_helper_dict2 = build_period_lookup_dict(input_dict = core_input_dict, input_dict_key_list=['thk_begin_date','thk_end_datetime'],time_period_choice_num='2')
    else:
        period_helper_dict2 = build_period_lookup_dict()
        
    print('DICT 2:',period_helper_dict2)

    print('Choose time period for Readmission/Excess Days.')
    #READM/EDAC quarter values
    if input_instructions_choice == '1':
        period_helper_dict3 = build_period_lookup_dict(input_dict = core_input_dict, input_dict_key_list=['readm_edac_quarters'],time_period_choice_num='3')

    else:
        period_helper_dict3 = build_period_lookup_dict()
        
    print('DICT 3:',period_helper_dict3)
        
        
    if input_instructions_choice == '1':
        #get begin and end dates for core measures
        begin_dts_cm = core_input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH']['most_measures_begin_date']
        end_dts_cm = core_input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH']['most_measures_end_datetime']
        
    else: 
        
        begin_dts_cm = input("Core Measures:  Enter the beginning datetime (format:  'xx-01-xxxx')")
        end_dts_cm = input("Core Measures:  Enter the end datetime (format:  'xx-xx-xxxx 23:59:59')")
    
    
    if input_instructions_choice == '1': 
    # step 2:  Import cohort data from the Vizient documentation file.
        cohort_helper_df = gather_cohort_data(input_dict_string = core_input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH']['cohort_files_wd'])
    
    else:
        cohort_helper_df = gather_cohort_data()
    
    if input_instructions_choice == '1': 
        # step 3:  Import the Vizient template hyperlink file and generate a helper dictionary from it.
        hyperlinks_helper_df = get_report_template_links_orig(input_dict_string = core_input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH']['template_hyperlink_wd'])
    else:
        hyperlinks_helper_df = get_report_template_links_orig()
   
    
    #join the hospital/hyperlink info to each hospital's cohort info
    merged_hyperlink_helper_df = pd.merge(cohort_helper_df, hyperlinks_helper_df[0], on='Hospital')
    
    #convert df to dictionary
    hyperlink_helper_dict_final = create_hyperlink_dict(merged_hyperlink_helper_df)

    
    if input_instructions_choice == '1': 
        # step 4:  Create main folder structure to store the files in.
        file_directory_name_helper = create_folder_structure(hyperlink_helper_dict_final, input_dict_string = core_input_dict['value_string_UPDATE_THIS_COLUMN_ONLY_EACH_MONTH']['cdb_report_output_wd'])
    else:
        file_directory_name_helper = create_folder_structure(hyperlink_helper_dict_final)
    
    
    # step 5:  Setup selenium driver and Login to Vizient main page
    browser_helper_obj = setup_webdriver_and_vizient_login()

    # step 6:  Loop over hyperlink helper dictionary and time period helper dictionary for every
    #          Vizient cohort and measure hyperlink, download the data, rename the file and store the excel file

    start = time.time()

    
    num_reports = loop_template_download_fy22(hyperlink_helper_dict_final, period_helper_dict1,period_helper_dict2,period_helper_dict3, browser_helper_obj,list_of_total_revisits_measures,file_directory_name_helper,hyperlinks_helper_df[1],remove_covid)
    
    end = time.time()
    elapsed_time = end - start
    
    return([num_reports, elapsed_time])
